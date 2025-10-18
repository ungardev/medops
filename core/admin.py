# Django core
from django.contrib import admin
from django.utils.html import format_html
from django.urls import path, reverse
from django.template.response import TemplateResponse
from django.utils.dateparse import parse_date
from django.http import HttpResponse
from django.db.models import Sum, Count, Avg
from datetime import datetime, date, timedelta
import matplotlib.pyplot as plt
from rangefilter.filters import DateRangeFilter, DateTimeRangeFilter, DateRangeFilterBuilder, NumericRangeFilter
from django.utils.timezone import now
from django.db.models.functions import TruncDate
import json
from django.core.serializers.json import DjangoJSONEncoder
from django.http import HttpRequest
from django.contrib.staticfiles import finders

# App models
from .models import (
    Patient,
    Appointment,
    Diagnosis,
    Treatment,
    Prescription,
    Payment,
    MedicalDocument,
)

# Librerías estándar
import csv
from io import BytesIO

# Excel
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# PDF
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from reportlab.lib import colors
from reportlab.platypus import (
    Table,
    TableStyle,
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Image,
)
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.utils import ImageReader


# Inline para documentos en Patient
class MedicalDocumentInlineForPatient(admin.TabularInline):
    model = MedicalDocument
    extra = 1
    fields = ("file", "description", "category", "preview_file")
    readonly_fields = ("preview_file",)

    @admin.display(description="Archivo")
    def preview_file(self, obj):
        if obj.file and obj.file.name.lower().endswith((".png", ".jpg", ".jpeg")):
            return format_html('<img src="{}" width="60" height="60" style="object-fit:cover;" />', obj.file.url)
        elif obj.file:
            return format_html('<a href="{}" target="_blank">Descargar</a>', obj.file.url)
        return "-"


# Inline para documentos en Appointment
class MedicalDocumentInlineForAppointment(admin.TabularInline):
    model = MedicalDocument
    extra = 1
    fields = ("file", "description", "category", "preview_file")
    readonly_fields = ("preview_file",)

    @admin.display(description="Archivo")
    def preview_file(self, obj):
        if obj.file and obj.file.name.lower().endswith((".png", ".jpg", ".jpeg")):
            return format_html('<img src="{}" width="60" height="60" style="object-fit:cover;" />', obj.file.url)
        elif obj.file:
            return format_html('<a href="{}" target="_blank">Descargar</a>', obj.file.url)
        return "-"


# Inline para documentos en Diagnosis
class MedicalDocumentInlineForDiagnosis(admin.TabularInline):
    model = MedicalDocument
    extra = 1
    fields = ("file", "description", "category", "preview_file")
    readonly_fields = ("preview_file",)

    @admin.display(description="Archivo")
    def preview_file(self, obj):
        if obj.file and obj.file.name.lower().endswith((".png", ".jpg", ".jpeg")):
            return format_html('<img src="{}" width="60" height="60" style="object-fit:cover;" />', obj.file.url)
        elif obj.file:
            return format_html('<a href="{}" target="_blank">Descargar</a>', obj.file.url)
        return "-"


@admin.register(Patient)
class PatientAdmin(admin.ModelAdmin):
    list_display = ('id', 'national_id', 'first_name', 'last_name', 'birthdate', 'gender', 'contact_info')
    list_display_links = ('id', 'national_id', 'first_name', 'last_name')
    search_fields = ('national_id', 'first_name', 'last_name', 'contact_info')
    ordering = ('last_name', 'first_name')
    list_per_page = 25
    inlines = [MedicalDocumentInlineForPatient]


@admin.register(Appointment)
class AppointmentAdmin(admin.ModelAdmin):
    list_display = ('id', 'patient', 'appointment_date', 'status')
    list_display_links = ('id', 'patient')
    list_filter = ('status', 'appointment_date')
    search_fields = ('patient__first_name', 'patient__last_name', 'patient__national_id')
    ordering = ('-appointment_date',)
    list_per_page = 25
    inlines = [MedicalDocumentInlineForAppointment]


@admin.register(Diagnosis)
class DiagnosisAdmin(admin.ModelAdmin):
    list_display = ('id', 'appointment', 'code', 'description')
    list_display_links = ('id', 'code')
    search_fields = ('code', 'description', 'appointment__patient__national_id')
    ordering = ('code',)
    list_per_page = 25
    inlines = [MedicalDocumentInlineForDiagnosis]


@admin.register(Treatment)
class TreatmentAdmin(admin.ModelAdmin):
    list_display = ('id', 'diagnosis', 'plan', 'start_date', 'end_date')
    list_display_links = ('id', 'diagnosis')
    ordering = ('-start_date',)
    list_per_page = 25


@admin.register(Prescription)
class PrescriptionAdmin(admin.ModelAdmin):
    list_display = ('id', 'diagnosis', 'medication', 'dosage', 'duration')
    list_display_links = ('id', 'medication')
    search_fields = ('medication', 'diagnosis__appointment__patient__national_id')
    ordering = ('medication',)
    list_per_page = 25


class AmountRangeFilter(admin.SimpleListFilter):
    title = "Rango de monto"
    parameter_name = "amount_range"

    def lookups(self, request, model_admin):
        return [
            ("low", "Menos de 100"),
            ("mid", "Entre 100 y 500"),
            ("high", "Más de 500"),
        ]

    def queryset(self, request, queryset):
        value = self.value()
        if value == "low":
            return queryset.filter(amount__lt=100)
        if value == "mid":
            return queryset.filter(amount__gte=100, amount__lte=500)
        if value == "high":
            return queryset.filter(amount__gt=500)
        return queryset


class MethodStatusFilter(admin.SimpleListFilter):
    title = "Método + Estado"
    parameter_name = "method_status"

    def lookups(self, request, model_admin):
        # Obtenemos los choices definidos en el modelo
        method_choices = model_admin.model._meta.get_field("method").choices
        status_choices = model_admin.model._meta.get_field("status").choices

        lookups = []
        for m_value, m_label in method_choices:
            for s_value, s_label in status_choices:
                key = f"{m_value}_{s_value}"
                label = f"{m_label} {s_label}"
                lookups.append((key, label))
        return lookups

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        try:
            method, status = value.split("_", 1)
            return queryset.filter(method=method, status=status)
        except ValueError:
            return queryset


class DateStatusFilter(admin.SimpleListFilter):
    title = "Fecha + Estado"
    parameter_name = "date_status"

    def lookups(self, request, model_admin):
        status_choices = model_admin.model._meta.get_field("status").choices
        lookups = []

        for s_value, s_label in status_choices:
            lookups.append((f"today_{s_value}", f"{s_label} hoy"))
            lookups.append((f"last7_{s_value}", f"{s_label} últimos 7 días"))
            lookups.append((f"this_month_{s_value}", f"{s_label} este mes"))
            lookups.append((f"this_year_{s_value}", f"{s_label} este año"))

        return lookups

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        today = date.today()
        try:
            period, status = value.split("_", 1)
        except ValueError:
            return queryset

        field = "appointment__appointment_date"

        if period == "today":
            start = datetime.combine(today, datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end), "status": status})

        if period == "last7":
            start = datetime.combine(today - timedelta(days=7), datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end), "status": status})

        if period == "this_month":
            start = datetime.combine(today.replace(day=1), datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end), "status": status})

        if period == "this_year":
            start = datetime.combine(date(today.year, 1, 1), datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end), "status": status})

        return queryset


class QuickDateRangeFilter(admin.SimpleListFilter):
    title = "Rango rápido"
    parameter_name = "quick_range"

    def lookups(self, request, model_admin):
        return [
            ("today", "Hoy"),
            ("last_7_days", "Últimos 7 días"),
            ("last_30_days", "Últimos 30 días"),
            ("this_month", "Este mes"),
            ("this_quarter", "Trimestre actual"),
            ("this_year", "Este año"),
        ]

    def queryset(self, request, queryset):
        value = self.value()
        if not value:
            return queryset

        field = "appointment__appointment_date"  # tu campo DateTimeField
        today = date.today()

        if value == "today":
            start = datetime.combine(today, datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end)})

        if value == "last_7_days":
            start = datetime.combine(today - timedelta(days=7), datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end)})

        if value == "last_30_days":
            start = datetime.combine(today - timedelta(days=30), datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end)})

        if value == "this_month":
            start = datetime.combine(today.replace(day=1), datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end)})

        if value == "this_quarter":
            quarter = (today.month - 1) // 3 + 1
            start_month = (quarter - 1) * 3 + 1
            start = datetime.combine(date(today.year, start_month, 1), datetime.min.time())

            if start_month + 2 == 12:
                end = datetime.combine(date(today.year, 12, 31), datetime.max.time())
            else:
                end = datetime.combine(date(today.year, start_month + 3, 1), datetime.min.time()) - timedelta(seconds=1)

            return queryset.filter(**{f"{field}__range": (start, end)})

        if value == "this_year":
            start = datetime.combine(date(today.year, 1, 1), datetime.min.time())
            end = datetime.combine(today, datetime.max.time())
            return queryset.filter(**{f"{field}__range": (start, end)})

        return queryset


@admin.register(Payment)
class PaymentAdmin(admin.ModelAdmin):
    list_display = (
        'id', 'patient_name', 'appointment', 'amount', 'method',
        'status', 'reference_number', 'bank_name', 'received_by', 'received_at'
    )
    list_display_links = ('id', 'appointment')

    list_filter = (
        'method',
        'status',
        ('appointment__appointment_date', DateTimeRangeFilter),
        QuickDateRangeFilter,
        ('amount', NumericRangeFilter),
        AmountRangeFilter,
        MethodStatusFilter,
        DateStatusFilter,  # 🔹 ahora con “últimos 7 días + estado”
    )

    search_fields = (
        'appointment__patient__first_name',
        'appointment__patient__last_name',
        'appointment__patient__national_id'
    )
    ordering = ('-appointment__appointment_date',)
    list_per_page = 25

    @admin.display(description="Paciente")
    def patient_name(self, obj):
        return f"{obj.appointment.patient.national_id} - {obj.appointment.patient.first_name} {obj.appointment.patient.last_name}"

    # 🔹 URLs personalizadas
    def get_urls(self):
        urls = super().get_urls()
        custom_urls = [
            path('report/', self.admin_site.admin_view(self.report_view), name="payment-report"),
            path('export-csv/', self.admin_site.admin_view(self.export_csv), name="payment-export-csv"),
            path('export-xlsx/', self.admin_site.admin_view(self.export_xlsx), name="payment-export-xlsx"),
            path('export-pdf/', self.admin_site.admin_view(self.export_pdf), name="payment-export-pdf"),
            path('dashboard/', self.admin_site.admin_view(self.dashboard_view), name="payment-dashboard"),  # 🔹 nuevo
        ]
        return custom_urls + urls
    
    def dashboard_view(self, request):
        method_data = list(
            Payment.objects.values("method")
            .annotate(total=Count("id"), amount=Sum("amount"))
            .order_by()
        )
        status_data = list(
            Payment.objects.values("status")
            .annotate(total=Count("id"), amount=Sum("amount"))
            .order_by()
        )
        timeline_data = list(
            Payment.objects.annotate(day=TruncDate("appointment__appointment_date"))
            .values("day")
            .annotate(total=Count("id"), amount=Sum("amount"))
            .order_by("day")
        )

        context = {
            **self.admin_site.each_context(request),
            "title": "Dashboard de Pagos",
            "method_data": method_data,
            "status_data": status_data,
            "timeline_data": timeline_data,
        }
        return TemplateResponse(request, "admin/payments_dashboard.html", context)
    
    # 🔹 Vista de reporte con filtros y resumen
    def report_view(self, request: HttpRequest):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        payments = Payment.objects.all()
        if start_date:
            payments = payments.filter(received_at__date__gte=parse_date(start_date))
        if end_date:
            payments = payments.filter(received_at__date__lte=parse_date(end_date))

        # 🔹 Querysets crudos para tablas
        totals_by_method_qs = (
            payments.values('method')
            .annotate(total_amount=Sum('amount'), count=Count('id'))
            .order_by('method')
        )

        totals_by_status_qs = (
            payments.values('status')
            .annotate(total_amount=Sum('amount'), count=Count('id'))
            .order_by('status')
        )

        # 🔹 Totales generales
        total_revenue = payments.aggregate(total=Sum('amount'))['total'] or 0
        total_payments = payments.count()
        avg_per_payment = payments.aggregate(avg=Avg('amount'))['avg'] or 0

        # 🔹 Contexto con doble salida
        context = dict(
            self.admin_site.each_context(request),
            # Para las tablas
            totals_by_method=totals_by_method_qs,
            totals_by_status=totals_by_status_qs,
            # Para los gráficos
            totals_by_method_json=json.dumps(list(totals_by_method_qs), cls=DjangoJSONEncoder),
            totals_by_status_json=json.dumps(list(totals_by_status_qs), cls=DjangoJSONEncoder),
            # Métricas globales
            title="Reporte Financiero de Pagos",
            start_date=start_date or "",
            end_date=end_date or "",
            total_revenue=total_revenue,
            total_payments=total_payments,
            avg_per_payment=avg_per_payment,
        )

        return TemplateResponse(request, "admin/payment_report.html", context)


    # 🔹 Exportador CSV
    def export_csv(self, request):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        payments = Payment.objects.all()
        if start_date:
            payments = payments.filter(received_at__date__gte=parse_date(start_date))
        if end_date:
            payments = payments.filter(received_at__date__lte=parse_date(end_date))

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="payments.csv"'

        writer = csv.writer(response)
        writer.writerow([
            "ID", "Paciente", "Cita", "Monto", "Método", "Estado",
            "Referencia", "Banco", "Recibido por", "Fecha registro"
        ])
        for p in payments:
            writer.writerow([
                p.pk,
                f"{p.appointment.patient.national_id} - {p.appointment.patient.first_name} {p.appointment.patient.last_name}",
                p.appointment.id,
                p.amount,
                p.method,
                p.status,
                p.reference_number or "",
                p.bank_name or "",
                p.received_by or "",
                p.received_at.strftime("%Y-%m-%d %H:%M") if p.received_at else ""
            ])
        return response

    # 🔹 Exportador Excel (XLSX) con hoja de resumen
    def export_xlsx(self, request):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        payments = Payment.objects.all()
        if start_date:
            payments = payments.filter(received_at__date__gte=parse_date(start_date))
        if end_date:
            payments = payments.filter(received_at__date__lte=parse_date(end_date))

        wb = Workbook()
        ws = wb.active  # type: ignore[attr-defined]
        ws.title = "Pagos"  # type: ignore[attr-defined]

        headers = [
            "ID", "Paciente", "Cita", "Monto", "Método", "Estado",
            "Referencia", "Banco", "Recibido por", "Fecha registro"
        ]
        ws.append(headers)  # type: ignore[attr-defined]

        # Encabezados en negrita
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)  # type: ignore[attr-defined]
            cell.font = Font(bold=True)

        # Filas de datos
        for p in payments:
            ws.append([  # type: ignore[attr-defined]
                p.pk,
                f"{p.appointment.patient.national_id} - {p.appointment.patient.first_name} {p.appointment.patient.last_name}",
                p.appointment.id,
                float(p.amount),
                p.method,
                p.status,
                p.reference_number or "",
                p.bank_name or "",
                p.received_by or "",
                p.received_at.strftime("%Y-%m-%d %H:%M") if p.received_at else ""
            ])

        # Ajustar ancho de columnas
        for col_idx in range(1, len(headers) + 1):
            max_length = 0
            for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=1, max_row=ws.max_row):  # type: ignore[attr-defined]
                for cell in col_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = max_length + 2  # type: ignore[attr-defined]

        # 🔹 Hoja de resumen
        ws_summary = wb.create_sheet(title="Resumen")  # type: ignore[attr-defined]

        # Totales por método
        totals_by_method = (
            payments.values('method')
            .annotate(total_amount=Sum('amount'), count=Count('id'))
            .order_by('method')
        )
        ws_summary.append(["Totales por Método"])  # type: ignore[attr-defined]
        ws_summary.append(["Método", "Cantidad", "Monto Total"])  # type: ignore[attr-defined]
        for row in totals_by_method:
            ws_summary.append([row['method'], row['count'], float(row['total_amount'] or 0)])  # type: ignore[attr-defined]
        ws_summary.append([])  # type: ignore[attr-defined]

        # Totales por estado
        totals_by_status = (
            payments.values('status')
            .annotate(total_amount=Sum('amount'), count=Count('id'))
            .order_by('status')
        )
        ws_summary.append(["Totales por Estado"])  # type: ignore[attr-defined]
        ws_summary.append(["Estado", "Cantidad", "Monto Total"])  # type: ignore[attr-defined]
        for row in totals_by_status:
            ws_summary.append([row['status'], row['count'], float(row['total_amount'] or 0)])  # type: ignore[attr-defined]
        ws_summary.append([])  # type: ignore[attr-defined]

        # Totales generales
        total_revenue = payments.aggregate(total=Sum('amount'))['total'] or 0
        total_payments = payments.count()
        avg_per_payment = payments.aggregate(avg=Avg('amount'))['avg'] or 0

        ws_summary.append(["Totales Generales"])  # type: ignore[attr-defined]
        ws_summary.append(["Total Recaudado", total_revenue])  # type: ignore[attr-defined]
        ws_summary.append(["Número de Pagos", total_payments])  # type: ignore[attr-defined]
        ws_summary.append(["Promedio por Pago", round(avg_per_payment, 2)])  # type: ignore[attr-defined]

        # Encabezados en negrita en la hoja de resumen
        for row in ws_summary.iter_rows(min_row=2, max_row=2, max_col=3):  # type: ignore[attr-defined]
            for cell in row:
                cell.font = Font(bold=True)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response['Content-Disposition'] = 'attachment; filename="payments.xlsx"'
        wb.save(response)  # type: ignore[attr-defined]
        return response
    
    # 🔹 Exportador PDF
    def export_pdf(self, request):
        start_date = request.GET.get("start_date")
        end_date = request.GET.get("end_date")

        queryset = Payment.objects.all()
        if start_date:
            queryset = queryset.filter(received_at__date__gte=parse_date(start_date))
        if end_date:
            queryset = queryset.filter(received_at__date__lte=parse_date(end_date))

        # Reutilizamos la versión avanzada
        return self.export_as_pdf(request, queryset)

    # 🔹 Botones extra en la lista de pagos
    def changelist_view(self, request, extra_context=None):
        if extra_context is None:
            extra_context = {}
        report_url = reverse("admin:payment-report")
        export_csv_url = reverse("admin:payment-export-csv")
        export_xlsx_url = reverse("admin:payment-export-xlsx")
        export_pdf_url = reverse("admin:payment-export-pdf")
        extra_context['report_button'] = format_html(
            '<a class="button" href="{}" style="margin-left:10px;">📊 Ver Reporte Financiero</a>'
            '<a class="button" href="{}" style="margin-left:10px;">⬇️ Exportar CSV</a>'
            '<a class="button" href="{}" style="margin-left:10px;">📑 Exportar Excel</a>'
            '<a class="button" href="{}" style="margin-left:10px;">🖨️ Exportar PDF</a>',
            report_url, export_csv_url, export_xlsx_url, export_pdf_url
        )
        return super().changelist_view(request, extra_context=extra_context)
    
    def export_as_xlsx(self, request):
        wb = Workbook()
        ws = wb.active  # type: ignore
        ws.title = "Pagos"  # type: ignore

        headers = ["ID", "Paciente", "Método", "Estado", "Monto", "Fecha"]
        ws.append(headers)  # type: ignore

        for p in Payment.objects.all():  # type: ignore
            created_at = getattr(p, "created_at", None)
            created_str = (
                created_at.strftime("%Y-%m-%d %H:%M") if created_at else ""
            )

            ws.append([  # type: ignore
                getattr(p, "id", ""),
                str(getattr(p, "patient", "")) if getattr(p, "patient", None) else "",
                getattr(p, "method", ""),
                getattr(p, "status", ""),
                float(getattr(p, "amount", 0)),
                created_str,
            ])

        for col_num, column_title in enumerate(headers, 1):
            col_letter = get_column_letter(col_num)
            ws.column_dimensions[col_letter].width = max(15, len(column_title) + 2)  # type: ignore

        output = BytesIO()
        wb.save(output)  # type: ignore
        output.seek(0)

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="payments.xlsx"'
        return response

    def export_as_pdf(self, request, queryset=None):
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)

        styles = getSampleStyleSheet()
        elements = []

        # 🔹 Logo y título
        logo_path = finders.find("core/img/medops-logo.png")  # ruta corregida
        if logo_path:
            from reportlab.lib.utils import ImageReader
            img = ImageReader(logo_path)
            iw, ih = img.getSize()
            aspect = ih / float(iw)
            # Escalar proporcionalmente a un ancho de 80 px
            logo = Image(logo_path, width=80, height=(80 * aspect))
        else:
            logo = Paragraph(" ", styles["Normal"])  # vacío si no hay logo

        title = Paragraph("Reporte Financiero de Pagos", styles["Title"])

        # 🔹 Encabezado: logo a la izquierda, título centrado
        header_table = Table([[logo, title]], colWidths=[100, 400])
        header_table.setStyle(TableStyle([
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("ALIGN", (0, 0), (0, 0), "LEFT"),     # logo alineado a la izquierda
            ("ALIGN", (1, 0), (1, 0), "CENTER"),   # título centrado
            ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ]))
        elements.append(header_table)
        elements.append(Spacer(1, 20))

        # 🔹 Encabezados de tabla principal
        headers = ["ID", "Paciente", "Método", "Estado", "Monto", "Fecha"]
        data = [headers]

        total_amount = 0.0
        method_totals = {}
        status_totals = {}

        payments = queryset if queryset is not None else Payment.objects.all()

        for payment in payments:
            created_at = getattr(payment, "received_at", None)
            created_str = created_at.strftime("%Y-%m-%d %H:%M") if created_at else ""

            amount = float(getattr(payment, "amount", 0))
            total_amount += amount

            method = getattr(payment, "method", "Desconocido")
            status = getattr(payment, "status", "Desconocido")

            method_totals[method] = method_totals.get(method, 0.0) + amount
            status_totals[status] = status_totals.get(status, 0.0) + amount

            row = [
                str(payment.pk),
                f"{payment.appointment.patient.first_name} {payment.appointment.patient.last_name}",
                method,
                status,
                f"{amount:.2f}",
                created_str,
            ]
            data.append(row)

        # 🔹 Totales generales
        data.append(["", "", "", "TOTAL", f"{total_amount:.2f}", ""])

        # 🔹 Tabla principal
        table = Table(data, colWidths=[40, 120, 80, 80, 80, 100])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#004080")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(table)

        # 🔹 Totales por método
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Totales por Método", styles["Heading2"]))
        method_data = [["Método", "Monto Total"]]
        for m, amt in method_totals.items():
            method_data.append([m, f"{amt:.2f}"])
        method_table = Table(method_data, colWidths=[150, 100])
        method_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#004080")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(method_table)

        # 🔹 Totales por estado
        elements.append(Spacer(1, 20))
        elements.append(Paragraph("Totales por Estado", styles["Heading2"]))
        status_data = [["Estado", "Monto Total"]]
        for s, amt in status_totals.items():
            status_data.append([s, f"{amt:.2f}"])
        status_table = Table(status_data, colWidths=[150, 100])
        status_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#004080")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ]))
        elements.append(status_table)

        # 🔹 Footer con número de página
        def add_page_number(canvas, doc):
            page_num = canvas.getPageNumber()
            text = f"Página {page_num} | MedOps Clinical System"
            canvas.setFont("Helvetica", 8)
            canvas.drawRightString(570, 20, text)

        # Construir documento con footer
        doc.build(elements, onFirstPage=add_page_number, onLaterPages=add_page_number)

        pdf = buffer.getvalue()
        buffer.close()

        response = HttpResponse(content_type="application/pdf")
        response["Content-Disposition"] = 'attachment; filename="payments_report.pdf"'
        response.write(pdf)
        return response
    
    # 🔹 Acción de exportación    
    @admin.action(description="Exportar pagos seleccionados a PDF")
    def export_selected_as_pdf(self, request, queryset):
        return self.export_as_pdf(request, queryset)


@admin.register(MedicalDocument)
class MedicalDocumentAdmin(admin.ModelAdmin):
    list_display = ('id', 'patient', 'appointment', 'diagnosis', 'description', 'category', 'uploaded_at', 'preview_file')
    list_display_links = ('id', 'description')
    list_filter = ('category', 'uploaded_at')
    search_fields = ('description', 'category', 'patient__first_name', 'patient__last_name', 'patient__national_id')
    ordering = ('-uploaded_at',)
    list_per_page = 25

    @admin.display(description="Archivo")
    def preview_file(self, obj):
        if obj.file and obj.file.name.lower().endswith((".png", ".jpg", ".jpeg")):
            return format_html('<img src="{}" width="80" height="80" style="object-fit:cover;" />', obj.file.url)
        elif obj.file:
            return format_html('<a href="{}" target="_blank">Descargar</a>', obj.file.url)
        return "-"

# Personalización del panel de administración
admin.site.site_header = "MedOps Clinical System"
admin.site.site_title = "MedOps Admin"
admin.site.index_title = "Panel de Control"
