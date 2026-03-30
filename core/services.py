import os
import io
from io import BytesIO
import re
import time
import base64
import hashlib
import logging
import tempfile
import traceback
import calendar
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from datetime import datetime, date, timedelta
from typing import Dict, Any, cast, Optional, List, Tuple, Union
#from PIL import Image as PILImage
from reportlab.platypus import Image as RLImage

# 2. Django Core
from django.conf import settings
from django.core.files import File
from django.core.files.base import ContentFile
from django.core.paginator import Paginator
from django.core.exceptions import ValidationError
from django.db import transaction
from django.db.models import Count, Sum, Q, F, Value, CharField
from django.db.models.functions import TruncDate, TruncMonth, TruncWeek, Coalesce, Cast, Concat
from django.http import JsonResponse, FileResponse, HttpResponse, HttpRequest
from django.shortcuts import get_object_or_404
from django.template.loader import render_to_string
from django.utils import timezone
from django.utils.dateparse import parse_date
from django.utils.timezone import now, localdate, make_aware

# 3. Herramientas de Terceros (Scraping, QR, Excel)
import requests
import qrcode
from bs4 import BeautifulSoup
from playwright.sync_api import sync_playwright
from weasyprint import HTML
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.drawing.image import Image as XLImage

# 4. ReportLab (Motor de PDF para Reportes Estructurados)
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)
from reportlab.lib.styles import getSampleStyleSheet

# 5. Django Rest Framework (DRF)
# Solo lo mínimo necesario para tipos en la capa de servicios
from rest_framework import status
from rest_framework.response import Response

# 6. Modelos Locales
from .models import (
    Patient, Appointment, Payment, Event, WaitingRoomEntry, GeneticPredisposition, 
    MedicalDocument, Diagnosis, Treatment, Prescription, ChargeOrder, ChargeItem, 
    InstitutionSettings, DoctorOperator, BCVRateCache, MedicalReport, ICD11Entry, 
    MedicalTest, MedicalReferral, Specialty, DocumentCategory, DocumentSource, 
    PersonalHistory, FamilyHistory, Surgery, Habit, Vaccine, VaccinationSchedule, 
    PatientVaccination, Allergy, MedicalHistory, ClinicalAlert, Country, State, 
    Municipality, City, Parish, Neighborhood, VitalSigns, ClinicalNote, DoctorPaymentConfig
)

# 7. Serializadores Locales
from .serializers import (
    PatientDetailSerializer, AppointmentDetailSerializer, DashboardSummarySerializer,
    PatientReadSerializer, PatientWriteSerializer, PatientListSerializer, 
    AppointmentSerializer, PaymentSerializer, WaitingRoomEntrySerializer, 
    WaitingRoomEntryDetailSerializer, GeneticPredispositionSerializer, 
    MedicalDocumentReadSerializer, MedicalDocumentWriteSerializer,
    AppointmentPendingSerializer, DiagnosisSerializer, TreatmentSerializer, 
    PrescriptionSerializer, ChargeOrderSerializer, ChargeItemSerializer, 
    ChargeOrderPaymentSerializer, EventSerializer, ReportRowSerializer, 
    ReportFiltersSerializer, ReportExportSerializer, InstitutionSettingsSerializer, 
    DoctorOperatorSerializer, MedicalReportSerializer, ICD11EntrySerializer, 
    DiagnosisWriteSerializer, MedicalTestSerializer, MedicalReferralSerializer, 
    PrescriptionWriteSerializer, TreatmentWriteSerializer, MedicalTestWriteSerializer, 
    MedicalReferralWriteSerializer, SpecialtySerializer, PersonalHistorySerializer, 
    FamilyHistorySerializer, SurgerySerializer, HabitSerializer, VaccineSerializer, 
    VaccinationScheduleSerializer, PatientVaccinationSerializer, 
    PatientClinicalProfileSerializer, AllergySerializer, MedicalHistorySerializer, 
    ClinicalAlertSerializer, CountrySerializer, StateSerializer, MunicipalitySerializer,
    CitySerializer, ParishSerializer, NeighborhoodSerializer, VitalSignsSerializer
)

# 8. Utils y Choices Locales
from core.utils.search_normalize import normalize, normalize_token
from .choices import UNIT_CHOICES, ROUTE_CHOICES, FREQUENCY_CHOICES

# SERVICIO OCR - Extracción de datos de pagos
import re
import logging
from typing import Optional, Dict, Any
from django.core.files.uploadedfile import UploadedFile


# Configuración de Logging
logger = logging.getLogger(__name__)


audit = logging.getLogger("audit")


def get_doctor_context() -> dict:
    doctor = DoctorOperator.objects.first()
    specialties = list(doctor.specialties.values_list("name", flat=True)) if doctor else []
    return {
        "full_name": doctor.full_name if doctor else "",
        "colegiado_id": doctor.colegiado_id if doctor else "",
        "specialties": specialties if specialties else ["No especificadas"],
        "signature": doctor.signature if (doctor and doctor.signature) else None,
    }

def get_patient_serialized(patient) -> dict:
    return dict(PatientDetailSerializer(patient).data)

def make_qr_data_uri(payload: str) -> str:
    img = qrcode.make(payload)
    buf = BytesIO()
    img.save(buf, "PNG")
    return f"data:image/png;base64,{base64.b64encode(buf.getvalue()).decode('utf-8')}"


def safe_json(value):
    return float(value) if isinstance(value, Decimal) else value


def generate_audit_code(appointment, patient):
    """
    Genera un código de auditoría único y trazable para documentos clínicos.
    Combina ID de consulta, ID de paciente y timestamp.
    """
    raw = f"{appointment.id}-{patient.id}-{timezone.now().isoformat()}"
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()[:12]  # 12 caracteres


def get_bcv_rate_logic():
    """
    SERVICIO PURO: Gestiona la obtención de la tasa BCV.
    Implementa: Cache -> Scraping Real -> Fallback al último valor.
    """
    today = timezone.localdate()
    
    # 1. Intentar obtener de la caché de hoy
    cache = BCVRateCache.objects.filter(date=today).first()
    if cache:
        return {
            "value": float(cache.value),
            "date": str(today),
            "source": "BCV_CACHE",
            "is_fallback": False
        }

    # 2. Si no hay caché, ejecutamos el Scraping con Playwright
    html = None
    try:
        with sync_playwright() as p:
            browser = p.chromium.launch(
                headless=True,
                args=["--disable-blink-features=AutomationControlled", "--no-sandbox"]
            )
            context = browser.new_context(
                user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
                locale="es-VE"
            )
            page = context.new_page()
            page.goto("https://www.bcv.org.ve/", wait_until="networkidle", timeout=30000)
            page.wait_for_selector("#dolar .centrado strong", timeout=15000)
            html = page.content()
            browser.close()
    except Exception as e:
        print(f"DEBUG: Error en Playwright: {str(e)}")

    # 3. Extraer y procesar el dato
    if html:
        soup = BeautifulSoup(html, "html.parser")
        dolar_elem = soup.select_one("#dolar .centrado strong")
        raw_val = dolar_elem.get_text(strip=True) if dolar_elem else None
        
        if raw_val:
            try:
                # Normalización venezolana: 36,45 -> 36.45
                normalized = raw_val.replace(".", "").replace(",", ".")
                rate_decimal = Decimal(normalized)
                
                # Guardar nuevo valor en caché
                BCVRateCache.objects.update_or_create(
                    date=today,
                    defaults={'value': rate_decimal}
                )
                
                return {
                    "value": float(rate_decimal),
                    "date": str(today),
                    "source": "BCV_LIVE",
                    "is_fallback": False
                }
            except Exception:
                pass

    # 4. ULTIMO RECURSO (Resiliencia Extrema)
    last_known = BCVRateCache.objects.order_by("-date").first()
    if last_known:
        return {
            "value": float(last_known.value),
            "date": str(last_known.date),
            "source": "BCV_HISTORY_FALLBACK",
            "is_fallback": True,
            "warning": "No se pudo conectar con el BCV, usando última tasa conocida"
        }
    
    raise RuntimeError("No hay conexión con el BCV ni datos históricos disponibles.")


def get_bcv_rate(): 
    """ 
    ✅ OPTIMIZADO: Siempre usa cache, nunca hace scraping en vivo.
    Mantiene compatibilidad con el resto del sistema devolviendo directamente un Decimal.
    
    Estrategia: 
    1. Buscar el rate más reciente en cache
    2. Si no existe, retornar 1.0 (fallback seguro)
    3. El scraping real debe hacerse via management command (scrape_bcv_rate)
    """
    try:
        # Intentar obtener el rate más reciente (de cualquier fecha)
        obj = BCVRateCache.objects.order_by("-date").first()
        if obj:
            return obj.value
    except Exception:
        pass
    
    # Fallback seguro si no hay cache
    return Decimal("1.0")


def get_dashboard_summary_data(start_date=None, end_date=None, range_param=None, currency="USD", status_param=None):
    """
    Músculo estadístico de Medopz.
    Calcula finanzas, citas y tendencias sin depender de la capa web.
    """
    today = localdate()

    # --- 1. Lógica de Fechas (Independiente) ---
    def parse_d(d):
        try:
            return parse_date(str(d)) if d else None
        except:
            return None

    if range_param == "day":
        start = end = today
    elif range_param == "week":
        start = today - timedelta(days=today.weekday())
        end = start + timedelta(days=6)
    elif range_param == "month":
        start = today.replace(day=1)
        end = date(today.year, today.month, calendar.monthrange(today.year, today.month)[1])
    else:
        start = parse_d(start_date) or (today - timedelta(days=6))
        end = parse_d(end_date) or today

    # --- 2. Finanzas (Agregaciones) ---
    orders_qs = ChargeOrder.objects.exclude(status="void")
    total_amount = orders_qs.aggregate(s=Sum("total")).get("s") or Decimal("0")
    confirmed_amount = Payment.objects.filter(status="confirmed").aggregate(s=Sum("amount")).get("s") or Decimal("0")
    balance_due = orders_qs.aggregate(s=Sum("balance_due")).get("s") or Decimal("0")

    waived_qs = ChargeOrder.objects.filter(status="waived")
    
    # --- 3. Clínico-operativo ---
    appt_qs = Appointment.objects.filter(appointment_date__range=(start, end))
    if status_param:
        appt_qs = appt_qs.filter(status=status_param)

    # --- 4. Tendencias (QuerySets optimizados) ---
    appt_trend = list(
        Appointment.objects.filter(appointment_date__range=(start, end), status="completed")
        .annotate(date=TruncDate("appointment_date"))
        .values("date")
        .annotate(value=Count("id"))
        .order_by("date")
    )

    pay_trend = list(
        Payment.objects.filter(status="confirmed", received_at__date__range=(start, end))
        .annotate(date=TruncDate("received_at"))
        .values("date")
        .annotate(value=Sum("amount"))
        .order_by("date")
    )

    # --- 5. Tasa BCV (Usando el servicio interno si existe o el Cache) ---
    latest_rate = BCVRateCache.objects.order_by("-created_at").first()
    rate_val = float(latest_rate.value) if latest_rate else 1.0
    
    def convert(amount):
        return float(amount) * rate_val if currency == "VES" else float(amount)

    # --- Payload Final ---
    return {
        "total_patients": Patient.objects.count(),
        "total_appointments": appt_qs.count(),
        "completed_appointments": appt_qs.filter(status="completed").count(),
        "pending_appointments": appt_qs.filter(status="pending").count(),
        "active_consultations": appt_qs.filter(status="in_consultation").count(),
        "waiting_room_count": WaitingRoomEntry.objects.filter(arrival_time__date__range=(start, end), status="waiting").count(),
        "total_payments_amount": convert(confirmed_amount),
        "financial_balance": convert(max(total_amount - balance_due, Decimal("0"))),
        "appointments_trend": [{"date": str(r["date"]), "value": int(r["value"])} for r in appt_trend],
        "payments_trend": [{"date": str(r["date"]), "value": float(r["value"])} for r in pay_trend],
        "bcv_rate": {
            "value": rate_val,
            "unit": "VES_per_USD",
            "is_fallback": latest_rate is None
        }
    }


def get_daily_appointments() -> List[Dict[str, Any]]:
    today = localdate()
    appointments = (
        Appointment.objects
        .filter(appointment_date=today)
        .select_related("patient")
        .order_by("arrival_time")
    )
    return cast(List[Dict[str, Any]], AppointmentSerializer(appointments, many=True).data)


def get_current_consultation() -> Optional[Dict[str, Any]]:
    """
    Obtiene la consulta activa actual.
    Busca cualquier appointment con status "in_consultation",
    ordenando por started_at descendente (más reciente primero).
    """
    appointment = (
        Appointment.objects
        .filter(status="in_consultation")
        .select_related("patient")
        .prefetch_related("diagnoses__treatments", "diagnoses__prescriptions")
        .order_by("-started_at")
        .first()
    )
    if not appointment:
        return None
    return cast(Dict[str, Any], AppointmentDetailSerializer(appointment).data)


def get_waitingroom_groups_today() -> Dict[str, Any]:
    today = localdate()
    groups_by_status = (
        WaitingRoomEntry.objects.filter(arrival_time__date=today)
        .values("status").annotate(total=Count("id")).order_by("status")
    )
    groups_by_priority = (
        WaitingRoomEntry.objects.filter(arrival_time__date=today)
        .values("priority").annotate(total=Count("id")).order_by("priority")
    )
    return {
        "by_status": list(groups_by_status),
        "by_priority": list(groups_by_priority),
    }


def recalc_appointment_status(appointment: Appointment):
    expected = Decimal(appointment.expected_amount or 0)
    total_paid = Payment.objects.filter(
        appointment=appointment, status="paid"
    ).aggregate(total=Sum("amount"))["total"] or Decimal("0")

    if total_paid >= expected and expected > 0:
        appointment.status = "paid"
    else:
        appointment.status = "pending"
    appointment.save(update_fields=["status"])


# --- Funciones auxiliares para escalar imágenes ---
def scaled_image(path: str, max_width: int, max_height: int) -> RLImage:
    """Escala una imagen para ReportLab manteniendo la proporción."""
    img = RLImage(path)
    iw, ih = img.drawWidth, img.drawHeight
    scale = min(max_width / iw, max_height / ih)
    img.drawWidth = iw * scale
    img.drawHeight = ih * scale
    return img

def scaled_excel_image(path: str, max_width: int, max_height: int) -> XLImage:
    img = XLImage(path)
    iw, ih = img.width, img.height
    scale = min(max_width / iw, max_height / ih)
    img.width = int(iw * scale)
    img.height = int(ih * scale)
    return img


def export_institutional_report(
    data_serialized: List[Dict[str, Any]], 
    export_format: str, 
    filters: Any, 
    target_currency: str, 
    user_name: str
) -> Tuple[io.BytesIO, str, str]:
    """
    SERVICIO: Genera el archivo binario (PDF o Excel) para exportaciones institucionales.
    Garantiza un retorno de (buffer, content_type, filename).
    """
    inst = InstitutionSettings.objects.first()
    doc_op = DoctorOperator.objects.first()
    
    # 1. Obtener tasa si es necesario
    rate = Decimal("1.0")
    if target_currency == "VES":
        # Asegúrate de que get_bcv_rate esté accesible o importada
        from .services import get_bcv_rate
        rate = get_bcv_rate()

    # 2. Preparar especialidades
    specialty_str = ""
    if doc_op and hasattr(doc_op, "specialties"):
        try:
            specialty_str = ", ".join([str(s) for s in doc_op.specialties.all()])
        except Exception:
            specialty_str = "No especificadas"

    buffer = io.BytesIO()

    # --- LÓGICA PDF ---
    if export_format == "pdf":
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()

        if inst:
            elements.append(Paragraph(f"<b>{inst.name or ''}</b>", styles["Title"]))
            elements.append(Paragraph(f"Dirección: {inst.address or ''}", styles["Normal"]))
            elements.append(Paragraph(f"Tel: {inst.phone or ''} • RIF: {inst.tax_id or ''}", styles["Normal"]))
            elements.append(Spacer(1, 12))

        if doc_op:
            elements.append(Paragraph(
                f"Médico operador: {doc_op.full_name or ''} • Colegiado: {doc_op.colegiado_id or ''} • {specialty_str}",
                styles["Normal"]
            ))
            elements.append(Spacer(1, 8))

        elements.append(Paragraph("<b>Reporte Institucional</b>", styles["Heading2"]))
        elements.append(Paragraph(f"Filtros: {str(filters)}", styles["Normal"]))
        elements.append(Paragraph(f"Tasa aplicada: {rate} Bs/USD", styles["Italic"]))
        elements.append(Spacer(1, 12))

        table_data = [["ID", "Fecha", "Tipo", "Entidad", "Estado", "Monto", "Moneda"]]
        for r in data_serialized:
            raw_date = r.get("date")
            date_str = str(raw_date)[:10] if raw_date else ""
            
            amount_dec = Decimal(str(r.get("amount") or "0"))
            amount_val = (amount_dec * rate).quantize(Decimal("0.01"), ROUND_HALF_UP)

            table_data.append([
                str(r.get("id") or ""),
                date_str,
                str(r.get("type") or ""),
                str(r.get("entity") or ""),
                str(r.get("status") or ""),
                f"{float(amount_val):.2f}",
                target_currency,
            ])

        table = Table(table_data, hAlign="LEFT")
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#003366")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ]))
        elements.append(table)
        
        elements.append(Spacer(1, 24))
        elements.append(Paragraph(f"Generado por: {user_name}", styles["Normal"]))
        elements.append(Paragraph(f"Fecha: {timezone.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))

        doc.build(elements)
        buffer.seek(0)
        return buffer, "application/pdf", "reporte.pdf"

    # --- LÓGICA EXCEL ---
    elif export_format == "excel":
        wb = Workbook()
        # 🔹 CAST para que Pylance reconozca ws como Worksheet y no como None
        from openpyxl.worksheet.worksheet import Worksheet
        ws = cast(Worksheet, wb.active)
        
        ws.title = "Reporte"

        if inst:
            # Acceso directo por celda tipado
            ws["A1"] = str(inst.name)
            ws["A1"].font = Font(bold=True, size=14)
        
        ws.append(["ID", "Fecha", "Tipo", "Entidad", "Estado", "Monto", "Moneda"])
        
        for r in data_serialized:
            raw_date = r.get("date")
            amount_dec = Decimal(str(r.get("amount") or "0"))
            amount_val = float((amount_dec * rate).quantize(Decimal("0.01"), ROUND_HALF_UP))
            
            ws.append([
                str(r.get("id") or ""), 
                str(raw_date)[:10] if raw_date else "", 
                str(r.get("type") or ""),
                str(r.get("entity") or ""), 
                str(r.get("status") or ""), 
                amount_val, 
                target_currency
            ])

        wb.save(buffer)
        buffer.seek(0)
        return buffer, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "reporte.xlsx"

    # Si llegamos aquí sin retorno, lanzamos error para que la vista no reciba None
    raise ValueError(f"Formato de exportación no soportado: {export_format}")


def generate_pdf_from_html(html: str, filename: str = "informe.pdf") -> File:
    """
    Convierte HTML en PDF y retorna un archivo Django File listo para guardar en un FileField.
    - Usa archivo temporal seguro.
    - Compatible con MedicalDocument.file.
    """
    from weasyprint import HTML
    with tempfile.NamedTemporaryFile(suffix=".pdf", delete=False) as tmp:
        # Renderizar PDF desde HTML
        HTML(string=html).write_pdf(tmp.name)
        tmp.seek(0)
        # Retornar como File listo para guardar
        return File(tmp, name=filename)


def generate_pdf_document(category: str, queryset, appointment):
    patient = appointment.patient
    institution = InstitutionSettings.objects.first()
    doctor = DoctorOperator.objects.first()
    specialties = list(doctor.specialties.values_list("name", flat=True)) if doctor else []

    # Paciente serializado
    patient_serialized = dict(PatientDetailSerializer(patient).data)

    # Generar audit code y QR
    audit_code = generate_audit_code(appointment, patient)
    qr_payload = f"Consulta:{appointment.id}|Category:{category}|Audit:{audit_code}"
    qr_img = qrcode.make(qr_payload)
    buffer = BytesIO()
    qr_img.save(buffer, "PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    qr_code_url = f"data:image/png;base64,{qr_base64}"

    # Helper defensivo
    def safe(val, default=""):
        return val if val is not None else default

    # Normalización de items según categoría
    items = []
    lab_tests, image_tests = [], []

    if category == "prescription":
        # Usar constantes globales, no atributos inexistentes en Prescription
        def route_label(val): return dict(ROUTE_CHOICES).get(val, val)
        def freq_label(val): return dict(FREQUENCY_CHOICES).get(val, val)
        def unit_label(val): return dict(UNIT_CHOICES).get(val, val)

        for p in queryset:
            med_name = ""
            if getattr(p, "medication_catalog", None):
                med_name = getattr(p.medication_catalog, "name", "") or getattr(p.medication_catalog, "title", "")
            if not med_name:
                med_name = safe(getattr(p, "medication_text", None))

            components = []
            for c in p.components.all():
                components.append({
                    "substance": c.substance,
                    "dosage": c.dosage,
                    "unit": unit_label(safe(c.unit)),
                })

            items.append({
                "medication": med_name or "Medicamento no especificado",
                "components": components,
                "route": route_label(safe(p.route)),
                "frequency": freq_label(safe(p.frequency)),
                "duration": safe(p.duration),
            })

    elif category == "treatment":
        for t in queryset:
            items.append({
                "description": safe(getattr(t, "plan", None)) or safe(getattr(t, "description", None)),
                "notes": safe(getattr(t, "notes", None)),
            })

    elif category == "medical_test_order":
        def urgency_label(val): return dict(MedicalTest.URGENCY_CHOICES).get(val, val)
        def status_label(val): return dict(MedicalTest.STATUS_CHOICES).get(val, val)
        def type_label(val): return dict(MedicalTest.TEST_TYPE_CHOICES).get(val, val)
        for t in queryset:
            row = {
                "type": type_label(safe(t.test_type)),
                "description": safe(t.description, "Sin descripción"),
                "urgency": urgency_label(safe(t.urgency)),
                "status": status_label(safe(t.status)),
            }
            if t.test_type in ["blood_test", "urine_test", "stool_test", "microbiology_culture", "biopsy", "genetic_test"]:
                lab_tests.append(row)
            elif t.test_type in ["xray", "ultrasound", "ct_scan", "mri", "ecg"]:
                image_tests.append(row)

    elif category == "medical_referral":
        for r in queryset:
            spec_names = [s.name for s in r.specialties.all()] if hasattr(r, "specialties") else []
            items.append({
                "notes": safe(r.reason) or safe(getattr(r, "notes", None)),
                "referred_to": safe(r.referred_to),
                "urgency": getattr(r, "get_urgency_display", lambda: safe(r.urgency))(),
                "status": getattr(r, "get_status_display", lambda: safe(r.status))(),
                "specialties": spec_names or [],
            })

    # Mapear categoría a template real
    template_map = {
        "treatment": "documents/treatment.html",
        "prescription": "documents/prescription.html",
        "medical_test_order": "documents/medical_test_order.html",
        "medical_referral": "documents/medical_referral.html",
    }
    tpl = template_map.get(category)
    if not tpl:
        raise ValueError(f"No existe template para la categoría {category}")

    context = {
        "appointment": appointment,
        "patient": patient_serialized,
        "institution": institution,
        "doctor": {
            "full_name": doctor.full_name if doctor else "",
            "colegiado_id": doctor.colegiado_id if doctor else "",
            "specialties": specialties if specialties else ["No especificadas"],
            "signature": doctor.signature if (doctor and doctor.signature) else None,
        },
        "items": items,
        "lab_tests": lab_tests,
        "image_tests": image_tests,
        "generated_at": timezone.now(),
        "audit_code": audit_code,
        "qr_code_url": qr_code_url,
    }

    html_string = render_to_string(tpl, context)
    pdf_bytes = HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf()

    # ✅ Devuelve siempre un tuple (pdf_file, audit_code)
    return ContentFile(pdf_bytes or b"", name=f"{category}_{appointment.id}.pdf"), audit_code


def get_audit_logic(
    entity: Optional[str] = None,
    entity_id: Optional[Union[int, str]] = None,
    patient_id: Optional[Union[int, str]] = None,
    limit: Optional[int] = None,
    split_by_category: bool = False,
    filters: Optional[Dict[str, Any]] = None
) -> Any:
    """
    SERVICIO MAESTRO DE AUDITORÍA: Centraliza 7 métodos en uno solo.
    Devuelve datos puros (dicts/lists), sin Response ni Request.
    
    ✅ FIX: Ahora busca eventos relacionados con el paciente a través de:
    - Patient (entity_id == patient_id)
    - Appointment (appointment.patient_id == patient_id)
    - ChargeOrder (charge_order.patient_id == patient_id)
    - Payment (payment.appointment.patient_id == patient_id)
    - WaitingRoomEntry (waiting_room_entry.patient_id == patient_id)
    """
    from .serializers import EventSerializer
    from .models import Appointment, ChargeOrder, Payment, WaitingRoomEntry
    # 1. Base Query optimizada
    qs = Event.objects.all().order_by("-timestamp")
    # 2. Filtros de Identidad (Búsqueda mejorada por patient_id)
    if patient_id:
        patient_id_int = int(patient_id)
        
        # Obtener IDs de entidades relacionadas con este paciente
        related_entity_ids = []
        
        # A) El paciente mismo (entity=Patient, entity_id=patient_id)
        patient_event_ids = list(Event.objects.filter(
            entity='Patient', 
            entity_id=patient_id_int
        ).values_list('id', flat=True))
        related_entity_ids.extend(patient_event_ids)
        
        # B) Citas del paciente
        appointment_ids = list(Appointment.objects.filter(
            patient_id=patient_id_int
        ).values_list('id', flat=True))
        
        if appointment_ids:
            # Citas directas
            appointment_event_ids = list(Event.objects.filter(
                entity='Appointment',
                entity_id__in=appointment_ids
            ).values_list('id', flat=True))
            related_entity_ids.extend(appointment_event_ids)
            
            # C) Órdenes de cargo del paciente
            charge_order_ids = list(ChargeOrder.objects.filter(
                patient_id=patient_id_int
            ).values_list('id', flat=True))
            
            if charge_order_ids:
                charge_order_event_ids = list(Event.objects.filter(
                    entity='ChargeOrder',
                    entity_id__in=charge_order_ids
                ).values_list('id', flat=True))
                related_entity_ids.extend(charge_order_event_ids)
                
                # D) Pagos relacionados con esas órdenes de cargo
                payment_ids = list(Payment.objects.filter(
                    charge_order_id__in=charge_order_ids
                ).values_list('id', flat=True))
                
                if payment_ids:
                    payment_event_ids = list(Event.objects.filter(
                        entity='Payment',
                        entity_id__in=payment_ids
                    ).values_list('id', flat=True))
                    related_entity_ids.extend(payment_event_ids)
            
            # E) Entradas en sala de espera del paciente
            waiting_room_ids = list(WaitingRoomEntry.objects.filter(
                patient_id=patient_id_int
            ).values_list('id', flat=True))
            
            if waiting_room_ids:
                waiting_event_ids = list(Event.objects.filter(
                    entity='WaitingRoomEntry',
                    entity_id__in=waiting_room_ids
                ).values_list('id', flat=True))
                related_entity_ids.extend(waiting_event_ids)
        
        # F) Búsqueda legacy en metadata (por si hay eventos con patient_id en metadata)
        metadata_events = Event.objects.filter(
            metadata__contains={"patient_id": patient_id_int}
        ).values_list('id', flat=True)
        related_entity_ids.extend(list(metadata_events))
        
        # Aplicar filtro: cualquier evento relacionado con el paciente
        if related_entity_ids:
            # Eliminar duplicados
            unique_ids = list(set(related_entity_ids))
            qs = qs.filter(id__in=unique_ids)
        else:
            # Si no hay eventos relacionados, devolver vacío
            qs = qs.none()
    
    # 3. Filtros originales
    if entity:
        qs = qs.filter(entity=entity)
    if entity_id:
        qs = qs.filter(entity_id=entity_id)
    # 4. Filtros Dinámicos (vienen de request.GET)
    if filters:
        if filters.get("start_date"):
            qs = qs.filter(timestamp__date__gte=filters["start_date"])
        if filters.get("end_date"):
            qs = qs.filter(timestamp__date__lte=filters["end_date"])
        if filters.get("severity"):
            qs = qs.filter(severity=filters["severity"])
        if filters.get("actor"):
            qs = qs.filter(actor_name__icontains=filters["actor"])
    # 5. Lógica de Categorización (Para Dashboards o Resúmenes)
    if split_by_category:
        lim = limit or 10
        clinical = qs.filter(entity__in=["Prescription", "Treatment"])[:lim]
        financial = qs.filter(entity__in=["Payment", "ChargeOrder"])[:lim]
        general = qs.filter(entity__in=["Patient", "Appointment", "WaitingRoomEntry"])[:lim]
        
        return {
            "clinical_events": EventSerializer(clinical, many=True).data,
            "financial_events": EventSerializer(financial, many=True).data,
            "general_events": EventSerializer(general, many=True).data,
            "all_events": EventSerializer(qs[:30], many=True).data
        }
    # 6. Dashboard Estadístico (Totalizaciones)
    if filters and filters.get("dashboard_stats"):
        return {
            "total_events": Event.objects.count(),
            "by_entity": list(Event.objects.values("entity").annotate(total=Count("id"))),
            "by_action": list(Event.objects.values("action").annotate(total=Count("id"))),
        }
    # 7. Retorno Simple (con o sin límite)
    if limit:
        qs = qs[:limit]
    else:
        qs = qs[:50]  # Límite por defecto
        
    return cast(List[Dict[str, Any]], EventSerializer(qs, many=True).data)


def get_payment_summary() -> List[Dict[str, Any]]:
    return list(Payment.objects.values("method").annotate(total=Sum("amount")))

def get_waived_consultations():
    return Payment.objects.filter(status="waived")


def get_waitingroom_today_data():
    today = timezone.localdate()
    return WaitingRoomEntry.objects.filter(
        Q(appointment__appointment_date=today) | 
        Q(arrival_time__date=today) | 
        Q(created_at__date=today)
    ).select_related(
        "patient", 
        "appointment",
        "institution"  # ✅ AGREGADO
    ).order_by("order", "arrival_time")

def get_pending_appointments():
    # Mueve aquí la lógica de cálculo de saldo que estaba en la vista
    appointments = Appointment.objects.select_related("patient").prefetch_related("payments")
    pending = []
    for appt in appointments:
        expected = float(appt.expected_amount or 0)
        total_paid = sum(float(p.amount or 0) for p in appt.payments.all() if p.status == "confirmed")
        if expected > total_paid:
            pending.append(appt)
    return pending


def get_report_data(report_type: str, start: Optional[date], end: Optional[date]) -> List[Dict[str, Any]]:
    # Aquí mueves los bloques 'if report_type == "financial"...' 
    # Retornando la lista de diccionarios 'data' que armaste en la vista original.
    ...

def update_institution_settings(data: Dict[str, Any], user) -> InstitutionSettings:
    """Actualiza la configuración de la institución y registra el evento."""
    settings_obj, _ = InstitutionSettings.objects.get_or_create(id=1)
    
    for key, value in data.items():
        if hasattr(settings_obj, key):
            setattr(settings_obj, key, value)
    
    settings_obj.save()
    
    Event.objects.create(
        entity="InstitutionSettings",
        entity_id=settings_obj.id,
        action="update_settings",
        actor=str(user),
        severity="info"
    )
    return settings_obj


def create_medical_document(
    patient, appointment, file_content, filename, category, user, **kwargs
) -> MedicalDocument:
    """Encapsula la creación del registro MedicalDocument y el guardado en disco."""
    # 1. Calcular Hash
    sha256 = hashlib.sha256(file_content).hexdigest()
    
    # 2. Guardar archivo físicamente
    file_path = os.path.join("medical_documents", filename)
    full_path = os.path.join(settings.MEDIA_ROOT, file_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    with open(full_path, "wb") as f:
        f.write(file_content)

    # 3. Crear registro en BD
    return MedicalDocument.objects.create(
        patient=patient,
        appointment=appointment,
        file=File(open(full_path, "rb"), name=filename),
        checksum_sha256=sha256,
        uploaded_by=user,
        category=category,
        **kwargs
    )


def create_medical_document_from_pdf(
    patient, appointment, pdf_bytes, filename, category, user, 
    diagnosis=None, description=None, audit_code=None, origin_panel="consultation"
) -> MedicalDocument:
    """Servicio para persistir un PDF generado como MedicalDocument."""
    file_path = os.path.join("medical_documents", filename)
    full_path = os.path.join(settings.MEDIA_ROOT, file_path)
    os.makedirs(os.path.dirname(full_path), exist_ok=True)
    
    with open(full_path, "wb") as f:
        f.write(pdf_bytes)

    django_file = File(open(full_path, "rb"), name=filename)
    
    sha256 = hashlib.sha256()
    for chunk in django_file.chunks():
        sha256.update(chunk)

    doc = MedicalDocument.objects.create(
        patient=patient,
        appointment=appointment,
        diagnosis=diagnosis,
        description=description or f"Documento {category} generado automáticamente",
        category=category,
        source="system_generated",
        origin_panel=origin_panel,
        template_version="v1.1",
        generated_by=user,
        uploaded_by=user,
        file=django_file,
        mime_type="application/pdf",
        size_bytes=django_file.size,
        checksum_sha256=sha256.hexdigest(),
        audit_code=audit_code
    )
    
    Event.objects.create(
        entity="MedicalDocument",
        entity_id=doc.id,
        action=f"generate_{category}",
        actor=str(user),
        metadata={"appointment_id": appointment.id, "category": category},
        severity="info",
        notify=True
    )
    return doc

def build_referral_pdf(referral, user) -> Tuple[bytes, str]:
    """Genera los bytes del PDF de una referencia médica."""
    context = {
        "referral": referral,
        "diagnosis": referral.diagnosis,
        "appointment": referral.appointment,
        "patient": referral.appointment.patient,
        "doctor": user,
        "institution": InstitutionSettings.objects.first(),
    }
    html_string = render_to_string("pdf/referral.html", context)
    # Pylance fix: or b"" garantiza que siempre sea bytes
    pdf_bytes = HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf() or b""
    filename = f"referral_{referral.id}.pdf"
    return pdf_bytes, filename

def build_chargeorder_pdf(charge_order, user) -> Tuple[bytes, str, str]:
    """Genera bytes de PDF de orden de cobro, QR y código de auditoría."""
    charge_order.recalc_totals()
    charge_order.save()

    patient = charge_order.patient
    appointment = charge_order.appointment
    doctor = DoctorOperator.objects.first()
    institution = InstitutionSettings.objects.first()
    audit_code = generate_audit_code(appointment, patient)

    # Generar QR
    qr_payload = f"Consulta:{appointment.id}|ChargeOrder:{charge_order.id}|Audit:{audit_code}"
    qr_img = qrcode.make(qr_payload)
    buffer = BytesIO()
    qr_img.save(buffer, "PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    qr_code_url = f"data:image/png;base64,{qr_base64}"

    context = {
        "charge_order": charge_order,
        "patient": patient,
        "appointment": appointment,
        "doctor": doctor,
        "institution": institution,
        "audit_code": audit_code,
        "qr_code_url": qr_code_url,
        "generated_at": timezone.now(),
    }

    html_string = render_to_string("pdf/charge_order.html", context)
    pdf_bytes = HTML(string=html_string, base_url=settings.MEDIA_ROOT).write_pdf() or b""
    filename = f"chargeorder_{charge_order.id}.pdf"
    
    return pdf_bytes, filename, audit_code


def get_daily_metrics() -> Dict[str, Any]:
    """
    Calcula y retorna las métricas institucionales del día actual.
    """
    today = localdate()
    
    return {
        "totalPatients": Patient.objects.count(),
        "todayAppointments": Appointment.objects.filter(appointment_date=today).count(),
        "pendingPayments": Payment.objects.filter(status="pending").count(),
        "waivedConsultations": Payment.objects.filter(status="waived").count(),
        "appointmentStatusToday": list(
            Appointment.objects.filter(appointment_date=today)
            .values("status")
            .annotate(total=Count("id"))
        ),
        "paymentMethodsTotals": list(
            Payment.objects.values("method")
            .annotate(total=Sum("amount"))
        ),
    }


def check_patient_safety(patient_id: int) -> List[Dict[str, Any]]:
    """
    Analiza el perfil del paciente para generar alertas preventivas inmediatas.
    """
    alerts = []
    patient = Patient.objects.get(pk=patient_id)
    
    # 1. Verificación de Alergias Críticas
    allergies = patient.allergies.all()
    if allergies.exists():
        msg = f"Alergias detectadas: {', '.join([a.name for a in allergies])}."
        alerts.append({"type": "danger", "message": msg, "icon": "ExclamationTriangleIcon"})

    # 2. Alerta de Antecedentes Genéticos
    genetic_risks = patient.genetic_predispositions.all()
    if genetic_risks.exists():
        alerts.append({
            "type": "warning", 
            "message": f"Riesgo genético: {', '.join([g.name for g in genetic_risks])}",
            "icon": "FingerPrintIcon"
        })

    # 3. Alerta de Signos Vitales (última toma)
    last_vitals = VitalSigns.objects.filter(patient=patient).order_by('-created_at').first()
    if last_vitals and last_vitals.is_abnormal: # Asumiendo lógica en el modelo
        alerts.append({
            "type": "critical",
            "message": "Últimos signos vitales fuera de rango normal.",
            "icon": "HeartIcon"
        })

    return alerts

def lock_consultation_integrity(appointment_id: int):
    """
    Cierra la consulta y genera un sello de integridad médico-legal.
    """
    with transaction.atomic():
        note = ClinicalNote.objects.get(appointment_id=appointment_id)
        if not note.is_locked:
            note.is_locked = True
            note.locked_at = timezone.now()
            # Generar un hash único de la nota para prevenir alteraciones en DB
            content_block = f"{note.subjective}{note.objective}{note.analysis}{note.plan}"
            note.security_hash = hashlib.sha256(content_block.encode()).hexdigest()
            note.save()
            
            # Actualizar estado de la cita
            note.appointment.status = "completed"
            note.appointment.save()
    return note


def generate_generic_pdf(instance: Any, category: str, institution: InstitutionSettings) -> Tuple[bytes, str, str]:
    """
    Fábrica universal de PDFs médicos con QR de auditoría.
    Soporta: prescription, treatment, medical_referral, medical_test_order.
    
    ✅ FASE 1: Context completamente enriquecido con TODOS los datos necesarios.
    """
    # ========================================
    # 1. OBTENER DATOS BASE SEGÚN CATEGORÍA
    # ========================================
    
    # Obtener patient según el tipo de instancia
    if hasattr(instance, 'patient'):
        patient = instance.patient
    elif hasattr(instance, 'appointment'):
        patient = instance.appointment.patient
    else:
        patient = None
    
    # Obtener appointment según el tipo de instancia
    if hasattr(instance, 'appointment'):
        appointment = instance.appointment
    elif hasattr(instance, 'diagnosis') and hasattr(instance.diagnosis, 'appointment'):
        appointment = instance.diagnosis.appointment
    else:
        appointment = None
    
    doctor = appointment.doctor if appointment else None
    
    # ✅ CAMBIO: Usar la institución pasada como parámetro
    # institution = InstitutionSettings.objects.first()  # ❌ Antes
    # institution = institution  # ✅ Ahora (el parámetro de la función)
    
    # ========================================
    # 2. CÓDIGO DE AUDITORÍA Y QR
    # ========================================
    
    raw_code = f"{category}-{instance.id}-{timezone.now().timestamp()}"
    audit_code = hashlib.sha256(raw_code.encode()).hexdigest()[:12].upper()
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(f"VERIFY_DOC:{audit_code}")
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    
    buffer = BytesIO()
    img_qr.save(buffer, kind="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    
    # ========================================
    # 3. SELECCIÓN DE PLANTILLA
    # ========================================
    
    template_map = {
        'prescription': 'documents/prescription.html',
        'treatment': 'documents/treatment.html',
        'medical_referral': 'medical/documents/medical_referral.html',
        'medical_test_order': 'documents/medical_test_order.html',
        'medical_report': 'medical/documents/medical_report.html',
        'charge_order': 'medical/documents/charge_order.html',
    }
    template_path = template_map.get(category, 'pdf/generic_medical_doc.html')
    
    # ========================================
    # 4. HELPERS DE FORMATEO
    # ========================================
    
    def calculate_age(birth_date):
        if not birth_date:
            return None
        today = date.today()
        try:
            birth = birth_date if isinstance(birth_date, date) else datetime.strptime(str(birth_date), "%Y-%m-%d").date()
            age = today.year - birth.year - ((today.month, today.day) < (birth.month, birth.day))
            return age
        except:
            return None
    
    def format_gender(gender):
        gender_map = {
            'M': 'Masculino',
            'F': 'Femenino',
            'male': 'Masculino',
            'female': 'Femenino',
            'O': 'Otro',
            'Other': 'Otro',
            'other': 'Otro',
        }
        return gender_map.get(gender, gender or 'No especificado')
    
    def format_urgency(urgency):
        urgency_map = {
            'routine': 'RUTINA',
            'urgent': 'URGENTE',
            'stat': 'STAT (INMEDIATO)',
            'priority': 'PRIORIDAD',
        }
        return urgency_map.get(urgency, urgency or 'No especificado')
    
    def format_status(status):
        status_map = {
            'pending': 'PENDIENTE',
            'collected': 'RECOLECTADA',
            'in_process': 'EN PROCESO',
            'completed': 'COMPLETADO',
            'cancelled': 'CANCELADO',
            'issued': 'EMITIDA',
            'accepted': 'ACEPTADA',
            'rejected': 'RECHAZADA',
        }
        return status_map.get(status, status or 'No especificado')
    
    def format_route(route):
        route_map = {
            'oral': 'Oral',
            'iv': 'Intravenosa',
            'im': 'Intramuscular',
            'sc': 'Subcutánea',
            'topical': 'Tópica',
            'sublingual': 'Sublingual',
            'inhalation': 'Inhalación',
            'rectal': 'Rectal',
            'other': 'Otra',
        }
        return route_map.get(route, route or 'No especificada')
    
    def format_frequency(frequency):
        frequency_map = {
            'once_daily': 'Una vez al día',
            'bid': 'Dos veces al día',
            'tid': 'Tres veces al día',
            'qid': 'Cuatro veces al día',
            'q4h': 'Cada 4 horas',
            'q6h': 'Cada 6 horas',
            'q8h': 'Cada 8 horas',
            'q12h': 'Cada 12 horas',
            'q24h': 'Cada 24 horas',
            'qod': 'Cada dos días',
            'stat': 'Inmediato',
            'prn': 'Según necesidad',
            'hs': 'Al acostarse',
            'ac': 'Antes de comer',
            'pc': 'Después de comer',
            'achs': 'Antes de comer y al acostarse',
        }
        return frequency_map.get(frequency, frequency or 'No especificada')
    
    # ========================================
    # 5. CONSTRUIR PATIENT_DATA
    # ========================================
    
    patient_data = None
    if patient:
        full_name = patient.full_name or ""
        name_parts = full_name.split(None, 1)
        first_name = name_parts[0] if name_parts else ""
        last_name = name_parts[1] if len(name_parts) > 1 else ""
        
        patient_data = {
            "id": patient.id,
            "full_name": patient.full_name or "",
            "first_name": first_name,
            "last_name": last_name,
            "national_id": patient.national_id or "",
            "age": calculate_age(patient.birthdate),
            "gender": format_gender(getattr(patient, 'gender', None)),
            "phone": getattr(patient, 'phone_number', None) or "",
            "email": getattr(patient, 'email', None) or "",
            "birth_date": patient.birthdate,
        }
    
    # ========================================
    # 6. APPOINTMENT_DATA (para todas las categorías)
    # ========================================
    
    appointment_data = None
    if appointment:
        appointment_data = {
            "id": appointment.id,
            "appointment_date": appointment.appointment_date,
            "status": appointment.status,
        }
    
    # ========================================
    # 7. CONTEXT ESPECÍFICO POR CATEGORÍA
    # ========================================
    
    if category == 'medical_referral':
        # ========================================
        # REFERENCIA MÉDICA
        # ========================================
        
        # Especialidades del médico referente
        doctor_specialties = []
        if doctor and hasattr(doctor, 'specialty') and doctor.specialty:
            doctor_specialties = [doctor.specialty.name]
        elif doctor and hasattr(doctor, 'specialties'):
            doctor_specialties = [s.name for s in doctor.specialties.all()]
        
        referring_doctor_data = None
        if doctor:
            referring_doctor_data = {
                "id": doctor.id,
                "full_name": doctor.full_name or "",
                "colegiado_id": getattr(doctor, 'colegiado_id', None) or "",
                "specialties": doctor_specialties,
                "signature": getattr(doctor, 'signature', None),
            }
        
        # Especialidades requeridas
        required_specialties = []
        if hasattr(instance, 'specialties'):
            required_specialties = [s.name for s in instance.specialties.all()]
        
        # Doctor de destino
        referred_to_doctor_name = ""
        if hasattr(instance, 'referred_to_doctor') and instance.referred_to_doctor:
            referred_to_doctor_name = instance.referred_to_doctor.full_name
        elif hasattr(instance, 'referred_to_external') and instance.referred_to_external:
            referred_to_doctor_name = instance.referred_to_external
        
        # Diagnósticos relacionados con la referencia
        diagnoses_data = []
        if hasattr(instance, 'diagnosis') and instance.diagnosis:
            diag = instance.diagnosis
            diagnoses_data.append({
                "icd_code": getattr(diag, 'icd_code', ''),
                "title": getattr(diag, 'title', '') or getattr(diag, 'name', ''),
                "description": getattr(diag, 'description', '') or getattr(diag, 'notes', ''),
            })
        
        context = {
            "referral": instance,
            "patient": patient_data,
            "appointment": appointment_data,
            "referring_doctor": referring_doctor_data,
            "institution": institution, # ✅ Usa el parámetro
            "audit_code": audit_code,
            "qr_code_url": f"data:image/png;base64,{qr_base64}",
            "generated_at": timezone.now(),
            "required_specialties": required_specialties,
            "referred_to_doctor": referred_to_doctor_name,
            "referred_to_institution": getattr(instance, 'referred_to_institution', None) or "",
            "referred_to_contact": getattr(instance, 'referred_to_contact', None) or "",
            "urgency_display": format_urgency(getattr(instance, 'urgency', None)),
            "status_display": format_status(getattr(instance, 'status', None)),
            "diagnoses": diagnoses_data,
            "instructions": getattr(instance, 'clinical_summary', None) or getattr(instance, 'instructions', None) or "",
        }
    
    elif category == 'prescription':
        # ========================================
        # PRESCRIPCIÓN / RECETA
        # ========================================
        
        doctor_specialties = []
        if doctor and hasattr(doctor, 'specialty') and doctor.specialty:
            doctor_specialties = [doctor.specialty.name]
        elif doctor and hasattr(doctor, 'specialties'):
            doctor_specialties = [s.name for s in doctor.specialties.all()]
        
        prescribing_doctor_data = None
        if doctor:
            prescribing_doctor_data = {
                "id": doctor.id,
                "full_name": doctor.full_name or "",
                "colegiado_id": getattr(doctor, 'colegiado_id', None) or "",
                "specialties": doctor_specialties,
                "signature": getattr(doctor, 'signature', None),
                "is_verified": getattr(doctor, 'is_verified', False),
            }
        
        # Nombre del medicamento
        medication_name = ""
        if hasattr(instance, 'medication_catalog') and instance.medication_catalog:
            medication_name = instance.medication_catalog.name
        elif hasattr(instance, 'medication_text') and instance.medication_text:
            medication_name = instance.medication_text
        
        # Construir items para la plantilla
        items_data = []
        item = {
            "medication": medication_name,
            "route": format_route(getattr(instance, 'route', None)),
            "frequency": format_frequency(getattr(instance, 'frequency', None)),
            "duration": getattr(instance, 'duration', None) or "",
            "notes": getattr(instance, 'indications', None) or "",
            "components": [],
        }
        
        # Agregar componentes si existen
        if hasattr(instance, 'components'):
            for comp in instance.components.all():
                item["components"].append({
                    "substance": comp.substance,
                    "dosage": comp.dosage,
                    "unit": comp.unit,
                })
        
        items_data.append(item)
        
        context = {
            "prescription": instance,
            "patient": patient_data,
            "appointment": appointment_data,
            "doctor": prescribing_doctor_data,
            "institution": institution, # ✅ Usa el parámetro
            "audit_code": audit_code,
            "qr_code_url": f"data:image/png;base64,{qr_base64}",
            "generated_at": timezone.now(),
            "medication_name": medication_name,
            "items": items_data,
        }
    
    elif category == 'treatment':
        # ========================================
        # TRATAMIENTO
        # ========================================
        
        doctor_specialties = []
        if doctor and hasattr(doctor, 'specialty') and doctor.specialty:
            doctor_specialties = [doctor.specialty.name]
        elif doctor and hasattr(doctor, 'specialties'):
            doctor_specialties = [s.name for s in doctor.specialties.all()]
        
        treating_doctor_data = None
        if doctor:
            treating_doctor_data = {
                "id": doctor.id,
                "full_name": doctor.full_name or "",
                "colegiado_id": getattr(doctor, 'colegiado_id', None) or "",
                "specialties": doctor_specialties,
                "signature": getattr(doctor, 'signature', None),
            }
        
        # Construir items para la plantilla
        items_data = []
        items_data.append({
            "description": getattr(instance, 'plan', None) or getattr(instance, 'title', None) or "",
            "notes": getattr(instance, 'notes', None) or "",
        })
        
        context = {
            "treatment": instance,
            "patient": patient_data,
            "appointment": appointment_data,
            "doctor": treating_doctor_data,
            "institution": institution, # ✅ Usa el parámetro
            "audit_code": audit_code,
            "qr_code_url": f"data:image/png;base64,{qr_base64}",
            "generated_at": timezone.now(),
            "items": items_data,
        }
    
    elif category == 'medical_test_order':
        # ========================================
        # ORDEN DE EXÁMENES MÉDICOS
        # ========================================
        
        doctor_specialties = []
        if doctor and hasattr(doctor, 'specialty') and doctor.specialty:
            doctor_specialties = [doctor.specialty.name]
        elif doctor and hasattr(doctor, 'specialties'):
            doctor_specialties = [s.name for s in doctor.specialties.all()]
        
        ordering_doctor_data = None
        if doctor:
            ordering_doctor_data = {
                "id": doctor.id,
                "full_name": doctor.full_name or "",
                "colegiado_id": getattr(doctor, 'colegiado_id', None) or "",
                "specialties": doctor_specialties,
            }
        
        # Obtener display del tipo de examen
        test_type_display = instance.get_test_type_display() if hasattr(instance, 'get_test_type_display') else getattr(instance, 'test_type', '')
        
        # Clasificar exámenes en laboratorios e imágenes
        # Categorías típicas de laboratorio
        lab_prefixes = ['hemogram', 'glucose', 'hemoglobin', 'platelets', 'coag', 'blood', 
                       'lipid', 'renal', 'liver', 'electrolytes', 'thyroid', 'bone', 
                       'cardiac', 'tumor', 'iron', 'vitamin', 'folate', 'amylase', 
                       'uric', 'protein', 'bilirubin', 'creatinine', 'urinalysis', 
                       'urine', 'stool', 'culture', 'hiv', 'hepatitis', 'autoimmune',
                       'rheumatoid', 'anti', 'crp', 'esr', 'allergy', 'immunoglobulin',
                       'complement', 'cryoglobulin', 'cortisol', 'acth', 'growth',
                       'prolactin', 'lh', 'testosterone', 'estradiol', 'progesterone',
                       'dhea', 'insulin', 'peptide', 'drug', 'alcohol', 'heavy',
                       'therapeutic', 'pregnancy', 'sweat', 'mantoux', 'patch']
        
        # Separar en lab_tests e image_tests
        lab_tests = []
        image_tests = []
        
        test_item = {
            "type": test_type_display,
            "description": getattr(instance, 'description', None) or "",
            "urgency": format_urgency(getattr(instance, 'urgency', None)),
            "status": format_status(getattr(instance, 'status', None)),
        }
        
        # Clasificar según el tipo de examen
        test_type = getattr(instance, 'test_type', '').lower()
        if any(test_type.startswith(prefix) for prefix in lab_prefixes):
            lab_tests.append(test_item)
        else:
            image_tests.append(test_item)
        
        # Si no hay clasificación, agregar a lab_tests por defecto
        if not lab_tests and not image_tests:
            lab_tests.append(test_item)
        
        context = {
            "test_order": instance,
            "patient": patient_data,
            "appointment": appointment_data,
            "doctor": ordering_doctor_data,
            "institution": institution, # ✅ Usa el parámetro
            "audit_code": audit_code,
            "qr_code_url": f"data:image/png;base64,{qr_base64}",
            "generated_at": timezone.now(),
            "test_type_display": test_type_display,
            "urgency_display": format_urgency(getattr(instance, 'urgency', None)),
            "status_display": format_status(getattr(instance, 'status', None)),
            "lab_tests": lab_tests,
            "image_tests": image_tests,
        }
    
    else:
        # ========================================
        # CATEGORÍA GENÉRICA
        # ========================================
        context = {
            "data": instance,
            "patient": patient_data,
            "appointment": appointment_data,
            "doctor": doctor,
            "institution": institution, # ✅ Usa el parámetro
            "audit_code": audit_code,
            "qr_code_url": f"data:image/png;base64,{qr_base64}",
            "generated_at": timezone.now(),
        }
    
    # ========================================
    # 8. RENDERIZAR PDF
    # ========================================
    
    html_string = render_to_string(template_path, context)
    
    pdf_bytes = HTML(
        string=html_string, 
        base_url=settings.MEDIA_ROOT
    ).write_pdf() or b""
    
    filename = f"{category}_{instance.id}_{audit_code}.pdf"
    
    return pdf_bytes, filename, audit_code


def generate_prescription_bundle(prescriptions: List[Any], appointment) -> Tuple[bytes, str, str]:
    """
    ✅ NUEVO: Genera un PDF consolidado con todas las prescripciones.
    """
    from django.template.loader import render_to_string
    from django.conf import settings
    import qrcode
    from io import BytesIO
    import hashlib
    from django.utils import timezone
    from core.models import InstitutionSettings
    
    # Obtener datos base
    patient = appointment.patient
    doctor = appointment.doctor
    institution = InstitutionSettings.objects.first()
    
    # Código de auditoría
    raw_code = f"prescription_bundle-{appointment.id}-{timezone.now().timestamp()}"
    audit_code = hashlib.sha256(raw_code.encode()).hexdigest()[:12].upper()
    
    # QR
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(f"VERIFY_DOC:{audit_code}")
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img_qr.save(buffer, kind="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    
    # ✅ Construir lista de items con toda la información
    items_data = []
    for pres in prescriptions:
        # Obtener diagnóstico asociado
        diagnosis_code = ""
        diagnosis_title = ""
        if hasattr(pres, 'diagnosis') and pres.diagnosis:
            diagnosis_code = getattr(pres.diagnosis, 'icd_code', '') or ''
            diagnosis_title = getattr(pres.diagnosis, 'title', '') or getattr(pres.diagnosis, 'description', '') or ''
        
        # Obtener nombre del medicamento
        medication_name = ""
        if hasattr(pres, 'medication_catalog') and pres.medication_catalog:
            medication_name = pres.medication_catalog.name
        elif hasattr(pres, 'medication_text') and pres.medication_text:
            medication_name = pres.medication_text
        
        # Verificar si es controlado
        is_controlled = False
        if hasattr(pres, 'medication_catalog') and pres.medication_catalog:
            is_controlled = getattr(pres.medication_catalog, 'is_controlled', False)
        
        # Formatear campos
        route = pres.get_route_display() if hasattr(pres, 'get_route_display') else getattr(pres, 'route', '') or ''
        frequency = pres.get_frequency_display() if hasattr(pres, 'get_frequency_display') else getattr(pres, 'frequency', '') or ''
        
        # Componentes
        components = []
        if hasattr(pres, 'components'):
            for comp in pres.components.all():
                components.append({
                    "substance": comp.substance,
                    "dosage": comp.dosage,
                    "unit": comp.unit,
                })
        
        items_data.append({
            "medication": medication_name,
            "route": route,
            "frequency": frequency,
            "duration": getattr(pres, 'duration', '') or '',
            "notes": getattr(pres, 'indications', '') or '',
            "components": components,
            "diagnosis_code": diagnosis_code,
            "diagnosis_title": diagnosis_title,
            "is_controlled": is_controlled,
        })
    
    # Calcular edad
    def calculate_age(birth_date):
        if not birth_date:
            return None
        from datetime import date
        today = date.today()
        try:
            birth = birth_date if isinstance(birth_date, date) else datetime.strptime(str(birth_date), "%Y-%m-%d").date()
            return today.year - birth.year - ((today.month, today.day) < (birth.month, birth.day))
        except:
            return None
    
    def format_gender(g):
        return {'M': 'Masculino', 'F': 'Femenino', 'male': 'Masculino', 'female': 'Femenino'}.get(g, g or 'No especificado')
    
    # Doctor data
    doctor_specialties = []
    if doctor:
        if hasattr(doctor, 'specialty') and doctor.specialty:
            doctor_specialties = [doctor.specialty.name]
        elif hasattr(doctor, 'specialties'):
            doctor_specialties = [s.name for s in doctor.specialties.all()]
    
    doctor_data = {
        "full_name": doctor.full_name if doctor else "",
        "colegiado_id": getattr(doctor, 'colegiado_id', '') or '',
        "specialties": doctor_specialties,
        "signature": getattr(doctor, 'signature', None),
    }
    
    patient_data = {
        "full_name": patient.full_name or "",
        "national_id": patient.national_id or "",
        "age": calculate_age(patient.birthdate),
        "gender": format_gender(getattr(patient, 'gender', None)),
    }
    
    appointment_data = {
        "id": appointment.id,
    }
    
    # Renderizar plantilla
    context = {
        "patient": patient_data,
        "appointment": appointment_data,
        "doctor": doctor_data,
        "institution": institution,
        "audit_code": audit_code,
        "qr_code_url": f"data:image/png;base64,{qr_base64}",
        "generated_at": timezone.now(),
        "items": items_data,
    }
    
    html_content = render_to_string('documents/prescription_bundle.html', context)
    
    # Generar PDF con WeasyPrint
    from weasyprint import HTML
    pdf_bytes = HTML(string=html_content, base_url=settings.MEDIA_ROOT).write_pdf()
    
    filename = f"prescriptions_bundle_{appointment.id}_{audit_code}.pdf"
    
    return pdf_bytes, filename, audit_code


def generate_treatment_bundle(treatments: List[Any], appointment) -> Tuple[bytes, str, str]:
    """
    ✅ NUEVO: Genera un PDF consolidado con todos los tratamientos.
    """
    from django.template.loader import render_to_string
    from django.conf import settings
    import qrcode
    from io import BytesIO
    import hashlib
    from django.utils import timezone
    from core.models import InstitutionSettings
    
    # Obtener datos base
    patient = appointment.patient
    doctor = appointment.doctor
    institution = InstitutionSettings.objects.first()
    
    # Código de auditoría
    raw_code = f"treatment_bundle-{appointment.id}-{timezone.now().timestamp()}"
    audit_code = hashlib.sha256(raw_code.encode()).hexdigest()[:12].upper()
    
    # QR
    qr = qrcode.QRCode(box_size=10, border=2)
    qr.add_data(f"VERIFY_DOC:{audit_code}")
    qr.make(fit=True)
    img_qr = qr.make_image(fill_color="black", back_color="white")
    buffer = BytesIO()
    img_qr.save(buffer, kind="PNG")
    qr_base64 = base64.b64encode(buffer.getvalue()).decode("utf-8")
    
    # ✅ Construir lista de items
    items_data = []
    for t in treatments:
        # Obtener tipo de tratamiento
        treatment_type = ""
        if hasattr(t, 'get_treatment_type_display'):
            treatment_type = t.get_treatment_type_display()
        elif hasattr(t, 'treatment_type'):
            treatment_type = t.treatment_type or ''
        
        # Fechas
        start_date = ""
        end_date = ""
        if hasattr(t, 'start_date') and t.start_date:
            start_date = t.start_date.strftime("%d/%m/%Y")
        if hasattr(t, 'end_date') and t.end_date:
            end_date = t.end_date.strftime("%d/%m/%Y")
        
        # Estado
        status = ""
        if hasattr(t, 'get_status_display'):
            status = t.get_status_display()
        elif hasattr(t, 'status'):
            status = t.status or ''
        
        items_data.append({
            "description": t.plan or t.title or "Sin descripción",
            "notes": getattr(t, 'notes', '') or '',
            "treatment_type": treatment_type,
            "start_date": start_date,
            "end_date": end_date,
            "status": status,
        })
    
    # Calcular edad
    def calculate_age(birth_date):
        if not birth_date:
            return None
        from datetime import date
        today = date.today()
        try:
            birth = birth_date if isinstance(birth_date, date) else datetime.strptime(str(birth_date), "%Y-%m-%d").date()
            return today.year - birth.year - ((today.month, today.day) < (birth.month, birth.day))
        except:
            return None
    
    def format_gender(g):
        return {'M': 'Masculino', 'F': 'Femenino', 'male': 'Masculino', 'female': 'Femenino'}.get(g, g or 'No especificado')
    
    # Doctor data
    doctor_specialties = []
    if doctor:
        if hasattr(doctor, 'specialty') and doctor.specialty:
            doctor_specialties = [doctor.specialty.name]
        elif hasattr(doctor, 'specialties'):
            doctor_specialties = [s.name for s in doctor.specialties.all()]
    
    doctor_data = {
        "full_name": doctor.full_name if doctor else "",
        "colegiado_id": getattr(doctor, 'colegiado_id', '') or '',
        "specialties": doctor_specialties,
        "signature": getattr(doctor, 'signature', None),
    }
    
    patient_data = {
        "full_name": patient.full_name or "",
        "national_id": patient.national_id or "",
        "age": calculate_age(patient.birthdate),
        "gender": format_gender(getattr(patient, 'gender', None)),
    }
    
    appointment_data = {
        "id": appointment.id,
    }
    
    # Renderizar plantilla
    context = {
        "patient": patient_data,
        "appointment": appointment_data,
        "doctor": doctor_data,
        "institution": institution,
        "audit_code": audit_code,
        "qr_code_url": f"data:image/png;base64,{qr_base64}",
        "generated_at": timezone.now(),
        "items": items_data,
    }
    
    html_content = render_to_string('documents/treatment_bundle.html', context)
    
    # Generar PDF con WeasyPrint
    from weasyprint import HTML
    pdf_bytes = HTML(string=html_content, base_url=settings.MEDIA_ROOT).write_pdf()
    
    filename = f"treatments_bundle_{appointment.id}_{audit_code}.pdf"
    
    return pdf_bytes, filename, audit_code


def bulk_generate_appointment_docs(appointment, user) -> Dict[str, Any]:
    """
    Genera automáticamente todos los documentos PDF de una cita.
    Implementación ELITE: Consolida múltiples prescripciones/tratamientos en un solo documento.
    
    ✅ FASE 1: Documentos consolidados (bundle) para prescripciones y tratamientos.
    """
    generated_files = []
    errors = []
    
    from core.models import Prescription, Treatment
    
    prescriptions = Prescription.objects.filter(diagnosis__appointment=appointment).select_related(
        'medication_catalog', 'diagnosis'
    ).prefetch_related('components')
    
    treatments = Treatment.objects.filter(diagnosis__appointment=appointment).select_related('diagnosis')
    
    # ✅ NUEVO: Generador consolidado para prescripciones
    if prescriptions.exists():
        try:
            # ✅ LÓGICA ELITE: Si hay más de 1 prescripción, generar bundle
            prescription_list = list(prescriptions)
            
            if len(prescription_list) == 1:
                # Comportamiento original: 1 documento por prescripción
                pdf_bytes, filename, audit_code = generate_generic_pdf(prescription_list[0], 'prescription', appointment.institution)
                description = f"Receta: {prescription_list[0].medication_catalog.name if prescription_list[0].medication_catalog else prescription_list[0].medication_text or 'Medicamento'}"
            else:
                # ✅ NUEVO: Generar documento consolidado
                pdf_bytes, filename, audit_code = generate_prescription_bundle(prescription_list, appointment)
                description = f"Recetas Consolidadas ({len(prescription_list)} medicamentos)"
            
            # Crear documento en BD
            doc = MedicalDocument.objects.create(
                patient=appointment.patient,
                appointment=appointment,
                doctor=appointment.doctor,
                institution=appointment.institution,
                generated_by=user,
                category='prescription',
                audit_code=audit_code,
                origin_panel="bulk_generator",
                description=description,
            )
            doc.file.save(filename, ContentFile(pdf_bytes))
            
            generated_files.append({
                "id": doc.id,
                "category": "prescription",
                "title": filename,
                "filename": filename,
                "audit_code": audit_code,
                "file_url": doc.file.url if doc.file else None,
                "description": description,
            })
            
        except Exception as e:
            errors.append({
                "category": "prescription",
                "item_id": None,
                "error": str(e)
            })
    
    # ✅ NUEVO: Generador consolidado para tratamientos
    if treatments.exists():
        try:
            # ✅ LÓGICA ELITE: Si hay más de 1 tratamiento, generar bundle
            treatment_list = list(treatments)
            
            if len(treatment_list) == 1:
                # Comportamiento original: 1 documento por tratamiento
                pdf_bytes, filename, audit_code = generate_generic_pdf(treatment_list[0], 'treatment', appointment.institution)
                title = treatment_list[0].plan or treatment_list[0].title or "Tratamiento"
                description = f"Tratamiento: {title[:100]}"
            else:
                # ✅ NUEVO: Generar documento consolidado
                pdf_bytes, filename, audit_code = generate_treatment_bundle(treatment_list, appointment)
                description = f"Tratamientos Consolidados ({len(treatment_list)} planes)"
            
            # Crear documento en BD
            doc = MedicalDocument.objects.create(
                patient=appointment.patient,
                appointment=appointment,
                doctor=appointment.doctor,
                institution=appointment.institution,
                generated_by=user,
                category='treatment',
                audit_code=audit_code,
                origin_panel="bulk_generator",
                description=description,
            )
            doc.file.save(filename, ContentFile(pdf_bytes))
            
            generated_files.append({
                "id": doc.id,
                "category": "treatment",
                "title": filename,
                "filename": filename,
                "audit_code": audit_code,
                "file_url": doc.file.url if doc.file else None,
                "description": description,
            })
            
        except Exception as e:
            errors.append({
                "category": "treatment",
                "item_id": None,
                "error": str(e)
            })
    
    # ✅ Mantener otros generadores (referencias, exámenes) - sin cambios
    other_generators = {
        'medical_referral': appointment.referrals.all().select_related('diagnosis').prefetch_related('specialties'),
        'medical_test_order': appointment.medical_tests.all(),
    }
    
    for category, queryset in other_generators.items():
        for item in queryset:
            try:
                pdf_bytes, filename, audit_code = generate_generic_pdf(item, category, appointment.institution)
                
                # Descripciones específicas por categoría
                if category == 'medical_referral':
                    specialties = [s.name for s in item.specialties.all()[:3]] if hasattr(item, 'specialties') else []
                    referred_to = item.referred_to_doctor.full_name if hasattr(item, 'referred_to_doctor') and item.referred_to_doctor else ""
                    description = f"Referencia: {', '.join(specialties)}" if specialties else f"Referencia a: {referred_to}" if referred_to else "Referencia Médica"
                else:
                    test_type = item.get_test_type_display() if hasattr(item, 'get_test_type_display') else getattr(item, 'test_type', 'Examen')
                    description = f"Orden de Examen: {test_type}"
                
                doc = MedicalDocument.objects.create(
                    patient=appointment.patient,
                    appointment=appointment,
                    doctor=appointment.doctor,
                    institution=appointment.institution,
                    generated_by=user,
                    category=category,
                    audit_code=audit_code,
                    origin_panel="bulk_generator",
                    description=description,
                )
                doc.file.save(filename, ContentFile(pdf_bytes))
                
                generated_files.append({
                    "id": doc.id,
                    "category": category,
                    "title": filename,
                    "filename": filename,
                    "audit_code": audit_code,
                    "file_url": doc.file.url if doc.file else None,
                    "description": description,
                })
                
            except Exception as e:
                errors.append({
                    "category": category,
                    "item_id": getattr(item, 'id', None),
                    "error": str(e)
                })
    
    return {
        "status": "success" if not errors else "partial_success",
        "total_generated": len(generated_files),
        "documents": generated_files,
        "generated_files": generated_files,
        "skipped": [],
        "errors": errors if errors else [],
    }


def get_advanced_metrics() -> Dict[str, Any]:
    today = localdate()
    yesterday = today - timedelta(days=1)
    
    # Métricas actuales
    current_patients = Patient.objects.count()
    current_revenue = Payment.objects.filter(received_at__date=today).aggregate(Sum('amount'))['amount__sum'] or 0
    
    # Cálculo de tendencia (Growth %)
    prev_patients = Patient.objects.filter(created_at__date=yesterday).count()
    growth = ((current_patients - prev_patients) / prev_patients * 100) if prev_patients > 0 else 100

    return {
        "summary": {
            "totalPatients": current_patients,
            "todayRevenue": float(current_revenue),
            "patientGrowth": round(growth, 2),
            "pendingPayments": Payment.objects.filter(status="pending").count(),
        },
        "appointmentVolume": list(
            Appointment.objects.filter(appointment_date__range=[today - timedelta(days=7), today])
            .values('appointment_date')
            .annotate(count=Count('id'))
            .order_by('appointment_date')
        )
    }


# ==========================================
# SERVICIOS DE CONFIGURACIÓN
# ==========================================
def get_institution_settings(request=None, active_only=False):
    """
    Obtiene todas las instituciones del doctor actual.
    
    Parámetros:
        request: Request HTTP (opcional)
        active_only: Si es True, devuelve solo la institución activa (por implementar)
    
    Retorna:
        - Todas las instituciones (array) si active_only=False
        - Una sola institución activa (objeto) si active_only=True
        - {} si no hay doctor configurado
    """
    try:
        # Obtener doctor actual
        doctor = None
        if request and hasattr(request, 'user') and request.user.is_authenticated:
            doctor = getattr(request.user, 'doctor_profile', None)
        
        if not doctor:
            institution = InstitutionSettings.objects.first()
            if active_only:
                return cast(Dict[str, Any], InstitutionSettingsSerializer(institution).data) if institution else {}
            return [cast(Dict[str, Any], InstitutionSettingsSerializer(institution).data)] if institution else []
        
        # Obtener todas las instituciones del doctor
        institutions = doctor.institutions.all()
        
        if active_only:
            institution_id = None
            if request:
                institution_id = request.META.get('HTTP_X_INSTITUTION_ID')
            
            if institution_id:
                institution = institutions.filter(id=institution_id).first()
            elif doctor.active_institution:
                institution = doctor.active_institution
            else:
                institution = institutions.first()
            
            return cast(Dict[str, Any], InstitutionSettingsSerializer(institution).data) if institution else {}
        
        # Devolver todas las instituciones del doctor
        serializer = InstitutionSettingsSerializer(institutions, many=True)
        return cast(Dict[str, Any], serializer.data)
    
    except Exception as e:
        logger.error(f"Error getting institution settings: {e}")
        raise


def update_institution_settings_ext(
    data: Dict[str, Any],
    user,
    files: Optional[Dict[str, Any]] = None
) -> InstitutionSettings:
    """
    Actualiza la configuración de la institución.
    Soporta FormData (con archivos) o JSON (sin archivos).
    
    Args:
        data: Datos a actualizar
        user: Usuario Django autenticado
        files: Archivos (para FormData)
    
    Returns:
        InstitutionSettings actualizada
    
    Raises:
        ValidationError: Si el doctor no puede editar esa institución
    """
    try:
        settings_obj = InstitutionSettings.objects.get(id=1)
        
        # VERIFICACIÓN: El doctor debe tener esta institución
        if user and user.is_authenticated and hasattr(user, 'doctor_profile'):
            doctor = user.doctor_profile
            if settings_obj not in doctor.institutions.all():
                raise ValidationError("No tienes permiso para editar esta institución")
        
        for key, value in data.items():
            if key in ["neighborhood_id", "neighborhood"]:
                if value and value != "":
                    settings_obj.neighborhood_id = value
            elif hasattr(settings_obj, key):
                setattr(settings_obj, key, value)
        
        if files:
            for key, file in files.items():
                if hasattr(settings_obj, key):
                    setattr(settings_obj, key, file)
        
        # ✅ FIX: Línea 1210 - Agregar 'user and' antes de 'user.is_authenticated'
        if user and user.is_authenticated:
            settings_obj.updated_by = user
        
        settings_obj.save()
        
        # ✅ FIX: Línea 1215 - También necesita verificar que user existe
        username = user.username if user and user.is_authenticated else "Anonymous"
        logger.info(f"Institution {settings_obj.name} updated by {username}")
        return settings_obj
    
    except InstitutionSettings.DoesNotExist:
        logger.error("Institution not found")
        raise
    except ValidationError as e:
        logger.error(f"Validation error updating institution: {e}")
        raise
    except Exception as e:
        logger.error(f"Error updating institution: {e}")
        raise


def get_doctor_config(request: Optional[HttpRequest] = None) -> Optional[Dict[str, Any]]:
    """
    Obtiene la configuración del médico operador autenticado.
    """
    import logging
    logger = logging.getLogger(__name__)
    
    # DIAGNOSTICO: Log de autenticación
    is_auth = request and request.user and request.user.is_authenticated
    logger.info(f"[GET_DOCTOR_CONFIG] is_authenticated={is_auth}, user={request.user if request else None}")
    
    if is_auth and request:
        doctor = getattr(request.user, 'doctor_profile', None)
        logger.info(f"[GET_DOCTOR_CONFIG] doctor from user.doctor_profile: id={doctor.id if doctor else None}")
        if not doctor:
            return None
    else:
        doctor = DoctorOperator.objects.first()
        logger.info(f"[GET_DOCTOR_CONFIG] doctor from fallback first(): id={doctor.id if doctor else None}")
        if not doctor:
            return None
    
    data = cast(Dict[str, Any], DoctorOperatorSerializer(doctor).data)
    
    if doctor.active_institution:
        data['active_institution'] = InstitutionSettingsSerializer(doctor.active_institution).data
    else:
        data['active_institution'] = None
    
    try:
        payment_config = doctor.payment_config
        logger.info(f"[GET_DOCTOR_CONFIG] payment_config found: bank_name={payment_config.bank_name}")
        data['bank_name'] = payment_config.bank_name or ""
        data['bank_rif'] = payment_config.bank_rif or ""
        data['bank_phone'] = payment_config.bank_phone or ""
        data['bank_account'] = payment_config.bank_account or ""
    except Exception as e:
        logger.warning(f"[GET_DOCTOR_CONFIG] payment_config NOT found: {e}")
        data['bank_name'] = ""
        data['bank_rif'] = ""
        data['bank_phone'] = ""
        data['bank_account'] = ""
    
    return data


def update_doctor_config(
    data: Dict[str, Any], 
    user, 
    files: Optional[Dict[str, Any]] = None
) -> DoctorOperator:
    """
    Actualiza la configuración del médico operador.
    Soporta FormData (con signature) o JSON.
    Maneja relaciones M2M (specialties, institutions).
    Maneja campos bancarios (DoctorPaymentConfig separado).
    """
    # ✅ CORREGIDO: Obtener el doctor por user si existe, o crear uno nuevo
    try:
        if user and hasattr(user, 'doctor_profile'):
            doctor_obj = user.doctor_profile
        else:
            colegiado_id = data.get('colegiado_id')
            if colegiado_id:
                doctor_obj = DoctorOperator.objects.filter(colegiado_id=colegiado_id).first()
                if not doctor_obj:
                    doctor_obj = DoctorOperator.objects.create(
                        colegiado_id=colegiado_id,
                        full_name="Operador Por Defecto",
                        gender="M"
                    )
            else:
                doctor_obj = DoctorOperator.objects.first()
                if not doctor_obj:
                    doctor_obj = DoctorOperator.objects.create(
                        full_name="Operador Por Defecto",
                        gender="M",
                        colegiado_id="N/A"
                    )
    except DoctorOperator.DoesNotExist:
        doctor_obj = DoctorOperator.objects.create(
            full_name="Operador Por Defecto",
            gender="M",
            colegiado_id="N/A"
        )
    
    # ✅ CORREGIDO: Manejar specialty_ids que vienen como strings del FormData
    specialty_ids = data.getlist("specialty_ids") if hasattr(data, 'getlist') else data.get("specialty_ids")
    
    if specialty_ids is not None:
        if isinstance(specialty_ids, list):
            try:
                ids = [int(sid) for sid in specialty_ids if sid]
            except (ValueError, TypeError):
                ids = []
        elif isinstance(specialty_ids, str):
            try:
                ids = [int(specialty_ids)] if specialty_ids else []
            except ValueError:
                ids = []
        else:
            ids = []
        
        if ids:
            doctor_obj.specialties.set(ids)
        else:
            doctor_obj.specialties.clear()
    
    # ✅ NUEVO: Manejar campos bancarios (DoctorPaymentConfig)
    bank_fields = ['bank_name', 'bank_rif', 'bank_phone', 'bank_account']
    bank_data = {k: v for k, v in data.items() if k in bank_fields}
    
    if bank_data:
        from .models import DoctorPaymentConfig
        payment_config, created = DoctorPaymentConfig.objects.get_or_create(
            doctor=doctor_obj,
            defaults=bank_data
        )
        if not created:
            for key, value in bank_data.items():
                setattr(payment_config, key, value)
            payment_config.save()
    
    # Actualización de campos básicos del DoctorOperator (excluir bank_* y specialty_ids)
    excluded_fields = ['specialty_ids', 'bank_name', 'bank_rif', 'bank_phone', 'bank_account']
    for key, value in data.items():
        if key in excluded_fields:
            continue
        if hasattr(doctor_obj, key) and key not in ['specialties', 'institutions']:
            try:
                setattr(doctor_obj, key, value)
            except Exception:
                pass
    
    # Procesar archivos (signature)
    if files:
        for field_name, file_obj in files.items():
            if hasattr(doctor_obj, field_name):
                try:
                    setattr(doctor_obj, field_name, file_obj)
                except Exception:
                    pass
    
    # ✅ CORREGIDO: Save solo si hay cambios
    try:
        doctor_obj.save()
    except Exception as e:
        logger.error(f"Error guardando DoctorOperator: {e}")
    
    # Log de auditoría
    try:
        Event.objects.create(
            entity="DoctorOperator",
            entity_id=doctor_obj.id,
            action="update_config",
            actor=str(user) if user else "system",
            metadata={k: str(v) for k, v in data.items() if k != "signature"},
            severity="info"
        )
    except Exception:
        pass
    
    return doctor_obj


def create_institution_for_doctor(data: Dict[str, Any], doctor: DoctorOperator) -> InstitutionSettings:
    """
    Crea una nueva institución y la agrega automáticamente al doctor actual.
    
    Args:
        data: Datos de la institución a crear
        doctor: DoctorOperator al que se agrega la institución
    
    Returns:
        InstitutionSettings creada
    
    Raises:
        ValidationError: Si hay errores de validación
    """
    try:
        serializer = InstitutionSettingsSerializer(data=data)
        serializer.is_valid(raise_exception=True)
        
        # ✅ FIX: cast() explicit para resolver el error de tipo
        institution = cast(InstitutionSettings, serializer.save())
        
        doctor.institutions.add(institution)
        
        if doctor.institutions.count() == 1:
            doctor.active_institution = institution
            doctor.save()
        
        logger.info(f"Institution created for doctor {doctor.colegiado_id}: {institution.name}")
        return institution
    
    except ValidationError as e:
        logger.error(f"Validation error creating institution: {e}")
        raise
    except Exception as e:
        logger.error(f"Error creating institution: {e}")
        raise


def add_institution_to_doctor(institution_id: int, doctor: DoctorOperator) -> InstitutionSettings:
    """
    Agrega una institución existente al doctor.
    
    Args:
        institution_id: ID de la institución existente
        doctor: DoctorOperator al que se agrega la institución
    
    Returns:
        InstitutionSettings agregada
    
    Raises:
        ValidationError: Si el doctor ya tiene esa institución
    """
    try:
        institution = InstitutionSettings.objects.get(id=institution_id)
        
        if institution in doctor.institutions.all():
            raise ValidationError(f"El doctor {doctor.colegiado_id} ya tiene la institución {institution.name}")
        
        doctor.institutions.add(institution)
        
        if doctor.institutions.count() == 1:
            doctor.active_institution = institution
            doctor.save()
        
        logger.info(f"Institution {institution.name} added to doctor {doctor.colegiado_id}")
        return institution
    
    except InstitutionSettings.DoesNotExist:
        logger.error(f"Institution {institution_id} not found")
        raise
    except ValidationError as e:
        logger.error(f"Validation error adding institution: {e}")
        raise
    except Exception as e:
        logger.error(f"Error adding institution to doctor: {e}")
        raise


def delete_institution_from_doctor(institution_id: int, doctor: DoctorOperator) -> Dict[str, Any]:
    """
    Elimina una institución del doctor.
    NO borra la institución de la DB global, solo la relación.
    
    Args:
        institution_id: ID de la institución a eliminar
        doctor: DoctorOperator al que se le elimina la institución
    
    Returns:
        Dict con éxito de la operación
    
    Raises:
        ValidationError: Si el doctor no tiene esa institución
        ValidationError: Si intenta eliminar la última institución
    """
    try:
        institution = InstitutionSettings.objects.get(id=institution_id)
        
        if institution not in doctor.institutions.all():
            raise ValidationError(f"El doctor {doctor.colegiado_id} no tiene la institución {institution.name}")
        
        if doctor.institutions.count() == 1:
            raise ValidationError("El doctor debe tener al menos una institución")
        
        doctor.institutions.remove(institution)
        
        if doctor.active_institution == institution:
            remaining = doctor.institutions.first()
            if remaining:
                doctor.active_institution = remaining
                doctor.save()
            else:
                doctor.active_institution = None
                doctor.save()
        
        logger.info(f"Institution {institution.name} deleted from doctor {doctor.colegiado_id}")
        return {"success": True, "message": "Institución eliminada correctamente"}
    
    except InstitutionSettings.DoesNotExist:
        logger.error(f"Institution {institution_id} not found")
        raise
    except ValidationError as e:
        logger.error(f"Validation error deleting institution: {e}")
        raise
    except Exception as e:
        logger.error(f"Error deleting institution from doctor: {e}")
        raise


def set_active_institution(institution_id: int, doctor: DoctorOperator) -> InstitutionSettings:
    """
    Cambia la institución activa (predeterminada) del doctor.
    Guarda en base de datos para persistencia entre sesiones.
    
    Args:
        institution_id: ID de la institución a marcar como activa
        doctor: DoctorOperator al que se le cambia la institución activa
    
    Returns:
        InstitutionSettings activa
    
    Raises:
        ValidationError: Si el doctor no tiene esa institución
    """
    try:
        institution = InstitutionSettings.objects.get(id=institution_id)
        
        if institution not in doctor.institutions.all():
            raise ValidationError(f"El doctor {doctor.colegiado_id} no tiene la institución {institution.name}")
        
        doctor.active_institution = institution
        doctor.save()
        
        logger.info(f"Active institution set to {institution.name} for doctor {doctor.colegiado_id}")
        return institution
    
    except InstitutionSettings.DoesNotExist:
        logger.error(f"Institution {institution_id} not found")
        raise
    except ValidationError as e:
        logger.error(f"Validation error setting active institution: {e}")
        raise
    except Exception as e:
        logger.error(f"Error setting active institution: {e}")
        raise


class PaymentOCRService:
    """
    Procesa imágenes de pagos móviles venezolanos y extrae datos automáticamente.
    Versión 5.0 - FASE 4: Múltiples estrategias + Hora + Receptor + Validación.
    
    Soporta: 31 bancos venezolanos + variaciones OCR comunes.
    Campos extraídos: banco, referencia, monto, teléfono, cédula, fecha, hora, receptor.
    Estrategias: 3 preprocesamientos × 3 configuraciones Tesseract = 9 intentos.
    """
    
    # === PATRONES DE BANCOS VENEZOLANOS (31 bancos + variaciones OCR) ===
    BANCO_PATTERNS = [
        # Top 5 bancos más usados
        (r'm[e3]rc[a4]nt[i1]l', '0105'),
        (r'ban[e3]sc[o0]', '0134'),
        (r'(?:bdv|banco\s*de\s*v[e3]n[e3]z[uú][e3]l[a4])', '0102'),
        (r'(?:bbva|pr[o0]v[i1]nc[i1]al)', '0108'),
        (r'bancam[i1]ga', '0172'),
        
        # Bancos medianos
        (r'bancar[i1]b[e3]', '0114'),
        (r'[e3]xt[e3]ri[o0]r', '0115'),
        (r'car[o0]n[i1]', '0128'),
        (r's[o0]f[i1]tasa', '0137'),
        (r'plaza', '0138'),
        (r'bang[e3]nt[e3]', '0146'),
        (r'f[o0]nd[o0]\s*com[uú]n', '0151'),
        (r'100\s*%\s*banco', '0156'),
        (r'd[e3]lsur', '0157'),
        (r't[e3]s[o0]r[o0]', '0163'),
        (r'bancr[e3]c[e3]r', '0168'),
        (r'r\s*4\s*banco', '0169'),
        (r'banco\s*activo', '0171'),
        (r'[i1]nt[e3]rnacional\s*d[e3]\s*d[e3]sarr[o0]ll[o0]', '0173'),
        (r'banpl[uú]s', '0174'),
        (r'd[i1]g[i1]tal.*trabajad[o0]r[e3]s', '0175'),
        (r'banfanb', '0177'),
        (r'n\s*58', '0178'),
        (r'nacional\s*d[e3]\s*cr[eé]d[i1]t[o0]', '0191'),
        (r'v[e3]n[e3]z[o0]lano\s*d[e3]\s*cr[eé]d[i1]t[o0]', '0104'),
        
        # Bancos adicionales
        (r'b[i1]c[e3]nt[e3]nario', '0116'),
        (r'[o0]cc[i1]d[e3]nt[e3]', '0100'),
        (r's[o0]b[e3]ran[o0]', '0100'),
        (r'agr[i1]c[o0]la', '0100'),
        (r'm[i1]\s*banco', '0100'),
        (r'c[i1]ty\s*partn[e3]r', '0100'),
    ]
    
    @classmethod
    def extract_data(cls, image) -> Dict[str, Any]:
        """
        Extrae datos de pago con múltiples estrategias de preprocesamiento.
        Intenta 3 configuraciones × 3 estrategias = 9 combinaciones.
        Toma el resultado con mayor confianza.
        """
        try:
            try:
                import pytesseract
                from PIL import Image
            except ImportError:
                return {
                    'success': False,
                    'error': 'OCR no disponible. Instale pytesseract y pillow.'
                }
            
            image.seek(0)
            img = Image.open(image)
            
            if img.mode != 'RGB':
                img = img.convert('RGB')
            
            # Configuraciones de Tesseract
            configs = [
                r'--oem 3 --psm 6',   # Bloque uniforme (mejor para apps)
                r'--oem 3 --psm 4',   # Columna de texto
                r'--oem 3 --psm 3',   # Página completa
            ]
            
            # Estrategias de preprocesamiento
            strategies = [
                ('binarization', cls.preprocess_binarization),
                ('high_contrast', cls.preprocess_high_contrast),
                ('inverted', cls.preprocess_inverted),
            ]
            
            best_result = None
            best_confidence = 0.0
            
            for strategy_name, preprocess_fn in strategies:
                try:
                    processed_img = preprocess_fn(img.copy())
                    
                    for config in configs:
                        text = pytesseract.image_to_string(
                            processed_img, 
                            lang='spa+eng', 
                            config=config
                        )
                        text = cls._normalize_ocr_text(text)
                        data = cls._parse_payment_text(text)
                        confidence = cls._calculate_confidence(data)
                        
                        if confidence > best_confidence:
                            best_confidence = confidence
                            best_result = {
                                'success': True,
                                'data': data,
                                'raw_text': text[:500],
                                'confianza': confidence,
                                'strategy': strategy_name,
                            }
                        
                        # Si ya tenemos 95%+ de confianza, no seguir
                        if confidence >= 0.95:
                            break
                    
                    if best_confidence >= 0.95:
                        break
                        
                except Exception as e:
                    logger.warning(f"Estrategia {strategy_name} falló: {e}")
                    continue
            
            if best_result:
                return best_result
            
            # Fallback: estrategia básica
            return cls._extract_basic(img)
            
        except Exception as e:
            logger.error(f"Error en OCR: {str(e)}")
            return {
                'success': False,
                'error': str(e)
            }
    
    # === ESTRATEGIAS DE PREPROCESAMIENTO ===
    
    @classmethod
    def preprocess_binarization(cls, img):
        """Estrategia 1: Binarización adaptativa (mejor para mayoría de capturas)"""
        try:
            from PIL import ImageEnhance, ImageFilter, ImageOps, Image
            import numpy as np
            
            max_dimension = 1200
            if max(img.size) > max_dimension:
                ratio = max_dimension / max(img.size)
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img = img.resize(new_size, Image.LANCZOS)
            
            img_gray = img.convert('L')
            img_array = np.array(img_gray)
            threshold = max(128, np.mean(img_array) * 0.7)
            img_binary = img_gray.point(lambda x: 255 if x > threshold else 0)
            img_clean = img_binary.filter(ImageFilter.MedianFilter(size=3))
            img_sharp = img_clean.filter(ImageFilter.SHARPEN)
            
            img_array_final = np.array(img_sharp)
            if np.mean(img_array_final) < 128:
                img_sharp = ImageOps.invert(img_sharp)
            
            return img_sharp.convert('RGB')
        except Exception as e:
            logger.warning(f"Binarización falló: {e}")
            return img.convert('RGB')
    
    @classmethod
    def preprocess_high_contrast(cls, img):
        """Estrategia 2: Alto contraste (mejor para capturas oscuras con texto claro)"""
        try:
            from PIL import ImageEnhance, ImageFilter, Image
            
            max_dimension = 1200
            if max(img.size) > max_dimension:
                ratio = max_dimension / max(img.size)
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img = img.resize(new_size, Image.LANCZOS)
            
            img_gray = img.convert('L')
            enhancer = ImageEnhance.Contrast(img_gray)
            img_contrast = enhancer.enhance(3.0)
            img_sharp = img_contrast.filter(ImageFilter.SHARPEN)
            img_sharp = img_sharp.filter(ImageFilter.SHARPEN)
            
            return img_sharp.convert('RGB')
        except Exception as e:
            logger.warning(f"Alto contraste falló: {e}")
            return img.convert('RGB')
    
    @classmethod
    def preprocess_inverted(cls, img):
        """Estrategia 3: Fondo invertido (mejor para apps con texto blanco sobre fondo oscuro)"""
        try:
            from PIL import ImageFilter, ImageOps, Image
            import numpy as np
            
            max_dimension = 1200
            if max(img.size) > max_dimension:
                ratio = max_dimension / max(img.size)
                new_size = (int(img.width * ratio), int(img.height * ratio))
                img = img.resize(new_size, Image.LANCZOS)
            
            img_gray = img.convert('L')
            img_inverted = ImageOps.invert(img_gray)
            img_array = np.array(img_inverted)
            threshold = np.mean(img_array) * 0.8
            img_binary = img_inverted.point(lambda x: 255 if x > threshold else 0)
            img_clean = img_binary.filter(ImageFilter.MedianFilter(size=3))
            
            return img_clean.convert('RGB')
        except Exception as e:
            logger.warning(f"Inversión falló: {e}")
            return img.convert('RGB')
    
    @classmethod
    def _extract_basic(cls, img):
        """Fallback: preprocesamiento básico original"""
        try:
            import pytesseract
            from PIL import ImageEnhance, ImageFilter
            
            img_gray = img.convert('L')
            enhancer = ImageEnhance.Contrast(img_gray)
            img_gray = enhancer.enhance(2.0)
            img_gray = img_gray.filter(ImageFilter.SHARPEN)
            img_rgb = img_gray.convert('RGB')
            
            text = pytesseract.image_to_string(img_rgb, lang='spa+eng', config=r'--oem 3 --psm 6')
            text = cls._normalize_ocr_text(text)
            data = cls._parse_payment_text(text)
            
            return {
                'success': True,
                'data': data,
                'raw_text': text[:500],
                'confianza': cls._calculate_confidence(data),
                'strategy': 'basic_fallback',
            }
        except Exception as e:
            return {'success': False, 'error': str(e)}
    
    # === NORMALIZACIÓN DE TEXTO OCR ===
    
    @classmethod
    def _normalize_ocr_text(cls, text: str) -> str:
        """Normaliza errores comunes de OCR en capturas bancarias."""
        import re
        
        # Reemplazar O entre dígitos → 0
        text = re.sub(r'(?<=\d)O(?=\d)', '0', text)
        text = re.sub(r'(?<=\d)o(?=\d)', '0', text)
        
        # Reemplazar I o l entre dígitos → 1
        text = re.sub(r'(?<=\d)[Il](?=\d)', '1', text)
        
        # Reemplazar S entre dígitos → 5
        text = re.sub(r'(?<=\d)S(?=\d)', '5', text)
        text = re.sub(r'(?<=\d)s(?=\d)', '5', text)
        
        # Reemplazar B entre dígitos → 8
        text = re.sub(r'(?<=\d)B(?=\d)', '8', text)
        
        return text
    
    # === EXTRACCIÓN DE CAMPOS ===
    
    @classmethod
    def _parse_payment_text(cls, text: str) -> Dict[str, Any]:
        """Parsea el texto OCR y extrae todos los campos de pago (8 campos)."""
        text_upper = text.upper()
        
        referencia = cls._extract_referencia(text)
        banco = cls._extract_banco(text_upper)
        monto = cls._extract_monto(text, exclude_ref=referencia)
        telefono = cls._extract_telefono(text)
        cedula = cls._extract_national_id(text)
        fecha = cls._extract_fecha(text)
        hora = cls._extract_hora(text)
        receptor = cls._extract_receptor(text)
        
        return {
            'banco': banco,
            'monto': monto,
            'referencia': referencia,
            'telefono': telefono,
            'cedula': cedula,
            'fecha': fecha,
            'hora': hora,
            'receptor': receptor,
        }
    
    @classmethod
    def _extract_banco(cls, text: str) -> Optional[str]:
        """Extrae el código del banco desde el texto OCR."""
        for pattern, bank_code in cls.BANCO_PATTERNS:
            if re.search(pattern, text, re.IGNORECASE):
                return bank_code
        return None
    
    @classmethod
    def _extract_referencia(cls, text: str) -> Optional[str]:
        """Extrae número de referencia (6-20 dígitos)."""
        patterns = [
            r'[Rr]eferencia[:\s]*(\d{6,20})',
            r'[Rr][Ee][Ff][\s.:]*(\d{6,20})',
            r'[Cc][óo]digo[:\s]*(\d{6,20})',
            r'[Nn][úu]mero\s*de\s*[Oo]peraci[oó]n[:\s]*(\d{6,20})',
            r'[Nn]ro?\s*[Oo]p[:\s]*(\d{6,20})',
            r'[Cc]ontrol[:\s]*(\d{6,20})',
            r'[Pp]ago\s*[Mm][óo]vil.*?(\d{6,20})',
            r'[Tt]ransferencia.*?(\d{12,20})',
            r'\b(\d{8,20})\b',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                ref = re.sub(r'[\s\-]', '', match.group(1))
                if 6 <= len(ref) <= 20:
                    return ref
        return None
    
    @classmethod
    def _extract_monto(cls, text: str, exclude_ref: Optional[str] = None) -> Optional[str]:
        """Extrae monto con soporte para formato venezolano (1.234.567,89)."""
        patterns = [
            r'[Bb][Ss]\.?\s*([\d]{1,3}(?:\.[\d]{3})*(?:,[\d]{2})?)',
            r'[Bb][Ss]\.?\s*([\d]+)',
            r'[Mm]onto[:\s]*[Bb][Ss]?\s*([\d.,]+)',
            r'[Tt]otal[:\s]*[Bb][Ss]?\s*([\d.,]+)',
            r'([\d]{1,3}(?:\.[\d]{3})+(?:,[\d]{2})?)\s*[Bb][Ss]',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.MULTILINE | re.IGNORECASE)
            if match:
                monto_str = match.group(1)
                
                if ',' in monto_str and '.' in monto_str:
                    monto_str = monto_str.replace('.', '').replace(',', '.')
                elif '.' in monto_str and monto_str.count('.') == 1:
                    parts = monto_str.split('.')
                    if len(parts[1]) != 2:
                        monto_str = monto_str.replace('.', '')
                elif ',' in monto_str:
                    monto_str = monto_str.replace(',', '.')
                
                if exclude_ref and monto_str == exclude_ref:
                    continue
                
                try:
                    valor = float(monto_str)
                    if 100 <= valor <= 100000000000:
                        if valor == int(valor):
                            return str(int(valor))
                        return str(valor)
                except (ValueError, OverflowError):
                    pass
        return None
    
    @classmethod
    def _extract_telefono(cls, text: str) -> Optional[str]:
        """Extrae teléfono en formato venezolano."""
        patterns = [
            r'(04\d{2}[-\s]?\d{7})',
            r'(04\d{9})',
            r'(\+58[\s-]?4\d{2}[\s-]?\d{3}[\s-]?\d{4})',
            r'(04\d{2}\s\d{3}\s\d{4})',
            r'(\(04\d{2}\)\s*\d{7})',
            r'[Tt]el[eé]fono[:\s]*(04\d{9})',
            r'[Cc]uenta[:\s]*(04\d{9})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text)
            if match:
                telefono = re.sub(r'[^\d]', '', match.group(1))
                if len(telefono) == 11 and telefono.startswith('04'):
                    return f"{telefono[:4]}-{telefono[4:]}"
                elif len(telefono) == 13 and telefono.startswith('58'):
                    telefono = '0' + telefono[2:]
                    return f"{telefono[:4]}-{telefono[4:]}"
        return None
    
    @classmethod
    def _extract_national_id(cls, text: str) -> Optional[str]:
        """Extrae cédula/RIF en formato venezolano."""
        patterns = [
            r'[Cc][ée]dula\s*(?:de\s*[Ii]dentidad)?[:\s]*([VEJPG])\s*[-\s]*(\d{5,9})',
            r'[Cc][Ii]\.?\s*[Ii][Dd]\.?\s*[:\s]*([VEJPG])\s*[-\s]*(\d{5,9})',
            r'[Rr][Ii][Ff]\.?\s*[:\s]*([VEJPG])\s*[-\s]*(\d{5,9})',
            r'\b([VEJPG])\s*-\s*(\d{6,9})\b',
            r'\b([VEJPG])(\d{7,9})\b',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                tipo = match.group(1).upper()
                numero = re.sub(r'[^\d]', '', match.group(2))
                if 5 <= len(numero) <= 12:
                    return f"{tipo}-{numero}"
        return None
    
    @classmethod
    def _extract_fecha(cls, text: str) -> Optional[str]:
        """Extrae fecha de la transacción."""
        from datetime import datetime
        
        patterns = [
            (r'\b(\d{2}/\d{2}/\d{4})\b', '%d/%m/%Y'),
            (r'\b(\d{2}-\d{2}-\d{4})\b', '%d-%m-%Y'),
            (r'\b(\d{2}\.\d{2}\.\d{4})\b', '%d.%m.%Y'),
            (r'\b(\d{4}-\d{2}-\d{2})\b', '%Y-%m-%d'),
            (r'\b(\d{2}/\d{2}/\d{4})\s+\d{2}:\d{2}\b', '%d/%m/%Y'),
        ]
        
        for pattern, fmt in patterns:
            match = re.search(pattern, text)
            if match:
                fecha_str = match.group(1)
                try:
                    fecha = datetime.strptime(fecha_str, fmt)
                    return fecha.strftime('%Y-%m-%d')
                except ValueError:
                    continue
        return None
    
    @classmethod
    def _extract_hora(cls, text: str) -> Optional[str]:
        """
        NUEVO: Extrae hora de la transacción.
        
        Formatos: HH:MM AM/PM, HH:MM:SS AM/PM, HH:MM (24h)
        """
        patterns = [
            (r'\b(\d{1,2}:\d{2}\s*[AaPp][Mm])\b', '%I:%M %p'),
            (r'\b(\d{1,2}:\d{2}:\d{2}\s*[AaPp][Mm])\b', '%I:%M:%S %p'),
            (r'\b(\d{2}:\d{2})\b', '%H:%M'),
            (r'\b(\d{2}:\d{2}:\d{2})\b', '%H:%M:%S'),
        ]
        
        for pattern, fmt in patterns:
            match = re.search(pattern, text)
            if match:
                hora_str = match.group(1).strip()
                try:
                    from datetime import datetime
                    hora = datetime.strptime(hora_str.upper(), fmt)
                    return hora.strftime('%H:%M')
                except ValueError:
                    continue
        return None
    
    @classmethod
    def _extract_receptor(cls, text: str) -> Optional[str]:
        """
        NUEVO: Extrae el nombre del receptor del pago.
        
        En capturas de pago móvil, el receptor es la persona o empresa
        que recibe el dinero.
        """
        patterns = [
            r'[Bb]eneficiario[:\s]*([A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,40})',
            r'[Rr]ecibe[:\s]*([A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,40})',
            r'[Rr]eceptor[:\s]*([A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,40})',
            r'[Dd]estinatario[:\s]*([A-ZÁÉÍÓÚÑa-záéíóúñ\s]{3,40})',
            r'[VEJPG]-\d{6,9}\s*\n?\s*([A-ZÁÉÍÓÚÑ\s]{3,40})',
        ]
        
        for pattern in patterns:
            match = re.search(pattern, text, re.IGNORECASE | re.MULTILINE)
            if match:
                nombre = match.group(1).strip()
                nombre = re.sub(r'^[\d\s]+|[\d\s]+$', '', nombre).strip()
                if len(nombre) >= 3:
                    return nombre.title()
        return None
    
    @classmethod
    def _calculate_confidence(cls, data: Dict[str, Any]) -> float:
        """
        Calcula nivel de confianza ponderado (8 campos).
        
        Pesos basados en importancia para verificación de pago:
        - referencia: 0.25 (identificador único de la transacción)
        - monto: 0.20 (cantidad transferida)
        - banco: 0.15 (app bancaria usada)
        - telefono: 0.10 (contacto del receptor)
        - cedula: 0.10 (identidad del receptor)
        - fecha: 0.08 (cuándo fue el pago)
        - hora: 0.07 (a qué hora)
        - receptor: 0.05 (nombre del receptor)
        """
        weights = {
            'referencia': 0.25,
            'monto': 0.20,
            'banco': 0.15,
            'telefono': 0.10,
            'cedula': 0.10,
            'fecha': 0.08,
            'hora': 0.07,
            'receptor': 0.05,
        }
        
        confidence = 0.0
        for field, weight in weights.items():
            if data.get(field) is not None and data[field] != '':
                confidence += weight
        
        return round(confidence, 2)